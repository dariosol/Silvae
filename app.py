#!/usr/bin/env python3
import sys, os, io, json, re, secrets, struct, tempfile, sqlite3, urllib.request
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from dotenv import load_dotenv
load_dotenv()  # loads .env when running locally; no-op on Railway

from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from geopy.geocoders import Nominatim
from datetime import datetime, timedelta, timezone
from sqlalchemy.exc import IntegrityError
from sqlalchemy import inspect, text, case
from werkzeug.security import generate_password_hash, check_password_hash
import jwt
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from functools import wraps
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from groq import Groq as GroqClient

from tools import dropdowns_ord as dd
from tools.lookup_tables import (
    species_lookup, monitoraggio as lt_monitoraggio, urgenza as lt_urgenza,
    conflitti as lt_conflitti, agenti_di_carie as lt_agenti_carie,
    altri_patogeni as lt_altri_patogeni,
    prescrizioni_valutative as lt_prescrizioni_val,
    prescrizioni_mitigazione as lt_prescrizioni_mit,
    prescrizioni_colturali as lt_prescrizioni_col,
)
from tools.ord_calculator import (
    assess_tree, bersaglio_value_to_class, bersaglio_flow_to_class,
    calc_ecological_value,
)
from tools import report_templates as report_tpl

# -----------------------
# Configuration
# -----------------------
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}})

DB_URL = os.environ.get('DATABASE_URL', 'postgresql://postgres:c0st4m4gn4@localhost/trees_db')
# Railway provides postgres:// — SQLAlchemy requires postgresql://
if DB_URL.startswith('postgres://'):
    DB_URL = DB_URL.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = DB_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
SECRET_KEY = os.environ.get('SECRET_KEY', 'super-secret-change-me')
app.config['SECRET_KEY'] = SECRET_KEY

SMTP_USER    = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
APP_BASE_URL = os.environ.get('APP_BASE_URL', 'http://localhost:5000')

# Demo accounts: shared read/write sandbox wiped and re-seeded to a fixed state
# on every demo login, so every visitor always sees the same demo of Torino.
# - demo_user (role "user"): the agronomer that owns the demo trees.
# - demo_city (role "city"): the municipality, linked to demo_user only, to
#   showcase the city role on the same trees.
DEMO_USERNAME       = os.environ.get('DEMO_USERNAME', 'demo_user')
DEMO_PASSWORD       = os.environ.get('DEMO_PASSWORD', 'demo_password')
DEMO_CITY_USERNAME  = os.environ.get('DEMO_CITY_USERNAME', 'demo_city')
DEMO_CITY_PASSWORD  = os.environ.get('DEMO_CITY_PASSWORD', 'demo_city_password')
DEMO_CITY_NAME      = 'Torino'

# Public self-registration. Disabled for now (not in production yet); re-enable by
# setting REGISTRATION_ENABLED=true (and un-hiding the "Crea account" link in the frontend).
REGISTRATION_ENABLED = os.environ.get('REGISTRATION_ENABLED', 'false').lower() == 'true'

db = SQLAlchemy(app)
geolocator = Nominatim(user_agent="tree_locator")

# -----------------------
# Models
# -----------------------
class User(db.Model):
    __tablename__ = 'user'
    id                     = db.Column(db.Integer, primary_key=True)
    username               = db.Column(db.String(80), unique=True, nullable=False)
    email                  = db.Column(db.String(120), unique=True, nullable=True)
    password_hash          = db.Column(db.String(255), nullable=False)
    role                   = db.Column(db.String(20), nullable=False)
    city                   = db.Column(db.String(100), nullable=True)
    password_reset_token   = db.Column(db.String(100), nullable=True)
    password_reset_expires = db.Column(db.DateTime(timezone=True), nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class CityMembership(db.Model):
    __tablename__ = 'city_membership'
    __table_args__ = (
        db.UniqueConstraint('city_user_id', 'agronomer_id', name='uq_city_agronomer'),
    )
    id           = db.Column(db.Integer, primary_key=True)
    city_user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    agronomer_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)

class City(db.Model):
    __tablename__ = 'city'
    id   = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), unique=True, nullable=False)

class Comune(db.Model):
    __tablename__ = 'comune'
    id        = db.Column(db.Integer, primary_key=True)
    nome      = db.Column(db.String(120), nullable=False, index=True)
    provincia = db.Column(db.String(100))
    sigla     = db.Column(db.String(2))
    regione   = db.Column(db.String(100))
    codice    = db.Column(db.String(10), unique=True)
    latitude  = db.Column(db.Float)
    longitude = db.Column(db.Float)

class Tree(db.Model):
    __tablename__ = 'tree'
    __table_args__ = ()
    # --- existing fields ---
    id               = db.Column(db.Integer, primary_key=True)
    custom_id        = db.Column(db.String(50), nullable=False)
    latitude         = db.Column(db.Float)
    longitude        = db.Column(db.Float)
    address          = db.Column(db.String(255))
    city             = db.Column(db.String(100), nullable=False)
    species          = db.Column(db.String(200), nullable=False)
    condition        = db.Column(db.String(50), nullable=False)
    comments         = db.Column(db.Text)
    actions          = db.Column(db.Text)
    height           = db.Column(db.String(50))
    trunk_diameter_cm= db.Column(db.Float)
    crown_diameter_m = db.Column(db.Float)
    age              = db.Column(db.String(50))
    location         = db.Column(db.String(255))
    cpc              = db.Column(db.String(50))
    next_check       = db.Column(db.Date)
    owner_id         = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    # --- ARETE dati generali ---
    dimora            = db.Column(db.String(100))
    stadio_sviluppo   = db.Column(db.String(50))
    posizione_sociale = db.Column(db.String(50))
    localizzazione    = db.Column(db.String(50))
    vincoli           = db.Column(db.String(100))

    # --- ORD dimensions (current) ---
    tree_height_m    = db.Column(db.Float)
    circonferenza_cm = db.Column(db.Float)
    branch_diam_cm   = db.Column(db.Float)
    branch_length_m  = db.Column(db.Float)
    branch_height_m  = db.Column(db.Float)
    target_height_m  = db.Column(db.Float)

    # --- ORD dimensions (post-intervention) ---
    post_tree_height_m    = db.Column(db.Float)
    post_circonferenza_cm = db.Column(db.Float)
    post_branch_diam_cm   = db.Column(db.Float)
    post_branch_length_m  = db.Column(db.Float)
    post_branch_height_m  = db.Column(db.Float)
    post_target_height_m  = db.Column(db.Float)

    # --- Pericolo (stored as string to allow '0 x sospet' etc.) ---
    pericolo_rami     = db.Column(db.String(20))
    pericolo_tronco   = db.Column(db.String(20))
    pericolo_colletto = db.Column(db.String(20))
    pericolo_zolla    = db.Column(db.String(20))

    # --- Bersaglio ---
    bersaglio_chioma_tipo  = db.Column(db.String(50))   # C32: proprietà/occupazione/pedoni/traffico…
    bersaglio_chioma_value = db.Column(db.String(100))  # description (proprietà/occupazione types)
    bersaglio_chioma_flow  = db.Column(db.Float)        # flow rate (pedoni/ora or auto/giorno)
    bersaglio_chioma       = db.Column(db.Integer)      # resolved class 1-7 (G32)
    bersaglio_ramo_tipo    = db.Column(db.String(50))   # N32
    bersaglio_ramo_value   = db.Column(db.String(100))
    bersaglio_ramo_flow    = db.Column(db.Float)        # flow rate for ramo
    bersaglio_ramo         = db.Column(db.Integer)      # resolved class 1-7 (R32)
    moltiplicatore         = db.Column(db.Integer)      # n simultaneous targets (H32); NULL = 1

    # --- Diagnosi (JSON: [{caratt, giudizio}, ...]) ---
    diag_zolla        = db.Column(db.Text)
    diag_colletto     = db.Column(db.Text)
    diag_fusto        = db.Column(db.Text)
    diag_castello     = db.Column(db.Text)
    diag_ramificazione= db.Column(db.Text)
    diag_chioma       = db.Column(db.Text)

    # --- Other multi-select (JSON arrays) ---
    conflitti_list   = db.Column(db.Text)
    agenti_carie     = db.Column(db.Text)
    altri_patogeni   = db.Column(db.Text)
    prescrizioni_val = db.Column(db.Text)
    prescrizioni_mit = db.Column(db.Text)
    prescrizioni_col = db.Column(db.Text)
    monitoraggio     = db.Column(db.String(50))
    urgenza          = db.Column(db.String(50))
    rischio          = db.Column(db.Text)   # JSON: latest assess_tree result

    # --- Valore ecologico (ORD row 10) ---
    condizione_salute_ecologica = db.Column(db.Text)   # dropdown input
    bio_kg                      = db.Column(db.Float)  # biomassa (kg)
    co2_kg_anno                 = db.Column(db.Float)  # CO2 sequestrata (kg/anno)
    o2_kg_anno                  = db.Column(db.Float)  # O2 prodotta (kg/anno)
    ia_kg_anno                  = db.Column(db.Float)  # intercett. acqua (kg/anno)
    valore_ecologico            = db.Column(db.Float)  # valore monetario (€)

class Inspection(db.Model):
    __tablename__ = 'inspection'
    id            = db.Column(db.Integer, primary_key=True)
    tree_id       = db.Column(db.Integer, db.ForeignKey('tree.id', ondelete='CASCADE'), nullable=False)
    date          = db.Column(db.Date, nullable=False)
    condition     = db.Column(db.String(50))
    comments      = db.Column(db.Text)
    actions       = db.Column(db.Text)
    inspector_id  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    inspector_name= db.Column(db.String(80))
    snapshot      = db.Column(db.Text)   # JSON snapshot of all tree fields at this moment
    rischio       = db.Column(db.Text)   # JSON: assess_tree result at inspection time
    created_at    = db.Column(db.DateTime(timezone=True),
                              default=lambda: datetime.now(timezone.utc))

# -----------------------
# DB init + migrate new columns
# -----------------------
with app.app_context():
    db.create_all()
    # Add any new columns that don't exist yet (safe for existing DBs)
    insp = inspect(db.engine)
    existing_cols = {c['name'] for c in insp.get_columns('tree')}
    new_cols = [
        ('dimora','VARCHAR(100)'), ('stadio_sviluppo','VARCHAR(50)'),
        ('posizione_sociale','VARCHAR(50)'), ('localizzazione','VARCHAR(50)'),
        ('vincoli','VARCHAR(100)'),
        ('tree_height_m','FLOAT'), ('circonferenza_cm','FLOAT'),
        ('branch_diam_cm','FLOAT'), ('branch_length_m','FLOAT'),
        ('branch_height_m','FLOAT'), ('target_height_m','FLOAT'),
        ('post_tree_height_m','FLOAT'), ('post_circonferenza_cm','FLOAT'),
        ('post_branch_diam_cm','FLOAT'), ('post_branch_length_m','FLOAT'),
        ('post_branch_height_m','FLOAT'), ('post_target_height_m','FLOAT'),
        ('pericolo_rami','VARCHAR(20)'), ('pericolo_tronco','VARCHAR(20)'),
        ('pericolo_colletto','VARCHAR(20)'), ('pericolo_zolla','VARCHAR(20)'),
        ('bersaglio_chioma_tipo','VARCHAR(50)'), ('bersaglio_chioma_value','VARCHAR(100)'),
        ('bersaglio_chioma_flow','FLOAT'), ('bersaglio_chioma','INTEGER'),
        ('bersaglio_ramo_tipo','VARCHAR(50)'), ('bersaglio_ramo_value','VARCHAR(100)'),
        ('bersaglio_ramo_flow','FLOAT'), ('bersaglio_ramo','INTEGER'),
        ('moltiplicatore','INTEGER'),
        ('diag_zolla','TEXT'), ('diag_colletto','TEXT'), ('diag_fusto','TEXT'),
        ('diag_castello','TEXT'), ('diag_ramificazione','TEXT'), ('diag_chioma','TEXT'),
        ('conflitti_list','TEXT'), ('agenti_carie','TEXT'), ('altri_patogeni','TEXT'),
        ('prescrizioni_val','TEXT'), ('prescrizioni_mit','TEXT'), ('prescrizioni_col','TEXT'),
        ('monitoraggio','VARCHAR(50)'), ('urgenza','VARCHAR(50)'),
        ('rischio','TEXT'),
        ('condizione_salute_ecologica','TEXT'),
        ('bio_kg','FLOAT'), ('co2_kg_anno','FLOAT'),
        ('o2_kg_anno','FLOAT'), ('ia_kg_anno','FLOAT'),
        ('valore_ecologico','FLOAT'),
    ]
    with db.engine.connect() as conn:
        for col_name, col_type in new_cols:
            if col_name not in existing_cols:
                conn.execute(text(f'ALTER TABLE tree ADD COLUMN {col_name} {col_type}'))
        conn.commit()
    existing_insp_cols = {c['name'] for c in insp.get_columns('inspection')}
    insp_new_cols = [('snapshot', 'TEXT'), ('rischio', 'TEXT')]
    with db.engine.connect() as conn:
        for col_name, col_type in insp_new_cols:
            if col_name not in existing_insp_cols:
                conn.execute(text(f'ALTER TABLE inspection ADD COLUMN {col_name} {col_type}'))
        conn.commit()
    existing_user_cols = {c['name'] for c in insp.get_columns('user')}
    user_new_cols = [
        ('email', 'VARCHAR(120)'),
        ('password_reset_token', 'VARCHAR(100)'),
        ('password_reset_expires', 'TIMESTAMPTZ'),
    ]
    with db.engine.connect() as conn:
        for col_name, col_type in user_new_cols:
            if col_name not in existing_user_cols:
                conn.execute(text(f'ALTER TABLE "user" ADD COLUMN {col_name} {col_type}'))
        conn.commit()
    with db.engine.connect() as conn:
        try:
            conn.execute(text('ALTER TABLE tree DROP CONSTRAINT uq_tree_custom_id_city'))
            conn.commit()
        except Exception:
            conn.rollback()
    existing_comune_cols = {c['name'] for c in insp.get_columns('comune')}
    comune_new_cols = [('latitude', 'DOUBLE PRECISION'), ('longitude', 'DOUBLE PRECISION')]
    with db.engine.connect() as conn:
        for col_name, col_type in comune_new_cols:
            if col_name not in existing_comune_cols:
                conn.execute(text(f'ALTER TABLE comune ADD COLUMN {col_name} {col_type}'))
        conn.commit()

# -----------------------
# Helpers
# -----------------------
def parse_json_field(v):
    if not v: return []
    try: return json.loads(v)
    except: return []

def tree_to_dict(t):
    return {
        'id': t.id, 'custom_id': t.custom_id,
        'latitude': t.latitude, 'longitude': t.longitude,
        'address': t.address, 'city': t.city,
        'species': t.species, 'condition': t.condition,
        'comments': t.comments, 'actions': t.actions,
        'height': t.height, 'trunk_diameter_cm': t.trunk_diameter_cm,
        'crown_diameter_m': t.crown_diameter_m, 'age': t.age,
        'location': t.location, 'cpc': t.cpc,
        'next_check': t.next_check.strftime("%Y-%m-%d") if t.next_check else None,
        'owner_id': t.owner_id,
        # ARETE
        'dimora': t.dimora, 'stadio_sviluppo': t.stadio_sviluppo,
        'posizione_sociale': t.posizione_sociale, 'localizzazione': t.localizzazione,
        'vincoli': t.vincoli,
        'tree_height_m': t.tree_height_m, 'circonferenza_cm': t.circonferenza_cm,
        'branch_diam_cm': t.branch_diam_cm, 'branch_length_m': t.branch_length_m,
        'branch_height_m': t.branch_height_m, 'target_height_m': t.target_height_m,
        'post_tree_height_m': t.post_tree_height_m,
        'post_circonferenza_cm': t.post_circonferenza_cm,
        'post_branch_diam_cm': t.post_branch_diam_cm,
        'post_branch_length_m': t.post_branch_length_m,
        'post_branch_height_m': t.post_branch_height_m,
        'post_target_height_m': t.post_target_height_m,
        'pericolo_rami': t.pericolo_rami, 'pericolo_tronco': t.pericolo_tronco,
        'pericolo_colletto': t.pericolo_colletto, 'pericolo_zolla': t.pericolo_zolla,
        'bersaglio_chioma_tipo':  t.bersaglio_chioma_tipo,
        'bersaglio_chioma_value': t.bersaglio_chioma_value,
        'bersaglio_chioma_flow':  t.bersaglio_chioma_flow,
        'bersaglio_chioma':       t.bersaglio_chioma,
        'bersaglio_ramo_tipo':    t.bersaglio_ramo_tipo,
        'bersaglio_ramo_value':   t.bersaglio_ramo_value,
        'bersaglio_ramo_flow':    t.bersaglio_ramo_flow,
        'bersaglio_ramo':         t.bersaglio_ramo,
        'moltiplicatore': t.moltiplicatore,
        'diag_zolla': parse_json_field(t.diag_zolla),
        'diag_colletto': parse_json_field(t.diag_colletto),
        'diag_fusto': parse_json_field(t.diag_fusto),
        'diag_castello': parse_json_field(t.diag_castello),
        'diag_ramificazione': parse_json_field(t.diag_ramificazione),
        'diag_chioma': parse_json_field(t.diag_chioma),
        'conflitti_list': parse_json_field(t.conflitti_list),
        'agenti_carie': parse_json_field(t.agenti_carie),
        'altri_patogeni': parse_json_field(t.altri_patogeni),
        'prescrizioni_val': parse_json_field(t.prescrizioni_val),
        'prescrizioni_mit': parse_json_field(t.prescrizioni_mit),
        'prescrizioni_col': parse_json_field(t.prescrizioni_col),
        'monitoraggio': t.monitoraggio, 'urgenza': t.urgenza,
        'rischio': json.loads(t.rischio) if t.rischio else None,
        'condizione_salute_ecologica': t.condizione_salute_ecologica,
        'bio_kg': t.bio_kg, 'co2_kg_anno': t.co2_kg_anno,
        'o2_kg_anno': t.o2_kg_anno, 'ia_kg_anno': t.ia_kg_anno,
        'valore_ecologico': t.valore_ecologico,
    }

def apply_tree_fields(tree_obj, data):
    """Set simple + JSON fields on a Tree instance from request data dict."""
    simple = [
        'species','condition','comments','actions','height','trunk_diameter_cm',
        'crown_diameter_m','age','location','cpc','address',
        'dimora','stadio_sviluppo','posizione_sociale','localizzazione','vincoli',
        'tree_height_m','circonferenza_cm','branch_diam_cm','branch_length_m',
        'branch_height_m','target_height_m',
        'post_tree_height_m','post_circonferenza_cm','post_branch_diam_cm',
        'post_branch_length_m','post_branch_height_m','post_target_height_m',
        'pericolo_rami','pericolo_tronco','pericolo_colletto','pericolo_zolla',
        'bersaglio_chioma_tipo','bersaglio_chioma_value','bersaglio_chioma_flow','bersaglio_chioma',
        'bersaglio_ramo_tipo','bersaglio_ramo_value','bersaglio_ramo_flow','bersaglio_ramo',
        'moltiplicatore',
        'monitoraggio','urgenza',
        'condizione_salute_ecologica',
    ]
    json_fields = [
        'diag_zolla','diag_colletto','diag_fusto','diag_castello',
        'diag_ramificazione','diag_chioma',
        'conflitti_list','agenti_carie','altri_patogeni',
        'prescrizioni_val','prescrizioni_mit','prescrizioni_col',
    ]
    for f in simple:
        if f in data:
            setattr(tree_obj, f, data[f])
    for f in json_fields:
        if f in data:
            v = data[f]
            setattr(tree_obj, f, json.dumps(v) if isinstance(v, (list, dict)) else v)
    # Auto-compute bersaglio class from tipo + value (proprietà/occupazione)
    # or tipo + flow rate (pedoni/ciclisti, traffico *).
    def _f(v):
        try: return float(v) if v not in (None, '') else None
        except: return None
    crown_diam = _f(data.get('crown_diameter_m')) or _f(getattr(tree_obj, 'crown_diameter_m', None)) or 0.0
    tree_h     = _f(data.get('tree_height_m'))    or _f(getattr(tree_obj, 'tree_height_m',    None)) or 0.0
    branch_l   = _f(data.get('branch_length_m'))  or _f(getattr(tree_obj, 'branch_length_m',  None)) or 0.0
    zone_chioma = (crown_diam + tree_h) / 2.0 if (crown_diam + tree_h) > 0 else 0.0
    zone_ramo   = branch_l * 1.25 if branch_l > 0 else 0.0

    for tipo_key, value_key, flow_key, class_key, zone_w in [
        ('bersaglio_chioma_tipo', 'bersaglio_chioma_value', 'bersaglio_chioma_flow',
         'bersaglio_chioma', zone_chioma),
        ('bersaglio_ramo_tipo',   'bersaglio_ramo_value',   'bersaglio_ramo_flow',
         'bersaglio_ramo',   zone_ramo),
    ]:
        tipo  = data.get(tipo_key)  or getattr(tree_obj, tipo_key,  None)
        value = data.get(value_key) or getattr(tree_obj, value_key, None)
        flow  = _f(data.get(flow_key)) or _f(getattr(tree_obj, flow_key, None))
        if tipo and value:
            computed = bersaglio_value_to_class(tipo, value)
            if computed is not None:
                setattr(tree_obj, class_key, computed)
        elif tipo and flow is not None and zone_w > 0:
            computed = bersaglio_flow_to_class(tipo, flow, zone_w)
            if computed is not None:
                setattr(tree_obj, class_key, computed)

    # Auto-compute valore ecologico when all required inputs are available
    eco = calc_ecological_value(
        trunk_diameter_cm          = _f(data.get('trunk_diameter_cm'))          or _f(getattr(tree_obj, 'trunk_diameter_cm', None)),
        tree_height_m              = _f(data.get('tree_height_m'))              or _f(getattr(tree_obj, 'tree_height_m', None)),
        stadio_sviluppo            = data.get('stadio_sviluppo')                or getattr(tree_obj, 'stadio_sviluppo', None),
        posizione_sociale          = data.get('posizione_sociale')              or getattr(tree_obj, 'posizione_sociale', None),
        condizione_salute_ecologica= data.get('condizione_salute_ecologica')    or getattr(tree_obj, 'condizione_salute_ecologica', None),
    )
    if eco is not None:
        tree_obj.bio_kg           = eco['bio_kg']
        tree_obj.co2_kg_anno      = eco['co2_kg_anno']
        tree_obj.o2_kg_anno       = eco['o2_kg_anno']
        tree_obj.ia_kg_anno       = eco['ia_kg_anno']
        tree_obj.valore_ecologico = eco['valore_ecologico']

def _calc_rischio(tree_obj):
    """Return assess_tree dict for tree_obj, or None if ORD fields are incomplete."""
    def _f(v):
        try: return float(v) if v not in (None, '') else None
        except: return None
    def _i(v):
        try: return int(v) if v not in (None, '') else None
        except: return None
    rq = {
        'tree_height_m':   _f(tree_obj.tree_height_m),
        'circumference_cm':_f(tree_obj.circonferenza_cm),
        'branch_diam_cm':  _f(tree_obj.branch_diam_cm),
        'branch_length_m': _f(tree_obj.branch_length_m),
        'branch_height_m': _f(tree_obj.branch_height_m),
        'target_height_m': _f(tree_obj.target_height_m),
        'pericolo_rami':   _i(tree_obj.pericolo_rami),
        'pericolo_tronco': _i(tree_obj.pericolo_tronco),
        'pericolo_colletto':_i(tree_obj.pericolo_colletto),
        'pericolo_zolla':  _i(tree_obj.pericolo_zolla),
        'bersaglio_chioma':_i(tree_obj.bersaglio_chioma),
        'bersaglio_ramo':  _i(tree_obj.bersaglio_ramo),
    }
    if any(v is None for v in rq.values()):
        return None
    mult = _i(tree_obj.moltiplicatore) or 1
    try:
        return assess_tree(
            crown_diam_m=_f(tree_obj.crown_diameter_m) or 0,
            moltiplicatore=mult,
            post_tree_height_m=_f(tree_obj.post_tree_height_m),
            post_circumference_cm=_f(tree_obj.post_circonferenza_cm),
            post_branch_diam_cm=_f(tree_obj.post_branch_diam_cm),
            post_branch_length_m=_f(tree_obj.post_branch_length_m),
            post_branch_height_m=_f(tree_obj.post_branch_height_m),
            post_target_height_m=_f(tree_obj.post_target_height_m),
            **rq
        )
    except Exception:
        return None

def _make_inspection(tree_obj, user_id, username):
    """Build an Inspection with snapshot + rischio for the current tree state."""
    rischio_val = _calc_rischio(tree_obj)
    snap = tree_to_dict(tree_obj)
    return Inspection(
        tree_id=tree_obj.id,
        date=datetime.now(timezone.utc).date(),
        condition=tree_obj.condition,
        comments=tree_obj.comments or '',
        actions=tree_obj.actions or '',
        inspector_id=user_id,
        inspector_name=username,
        snapshot=json.dumps(snap),
        rischio=json.dumps(rischio_val) if rischio_val else None,
    )

# -----------------------
# Demo account seeding
# -----------------------
# Fixed set of demo trees in Torino. Owned by the demo user, wiped and re-created
# on every demo login so the demo always looks identical. A few carry full ORD
# measures so the risk (B·I·P) is computed; others are left partial on purpose.
DEMO_TREES = [
    dict(custom_id='DEMO-001', species='Platanus × acerifolia', condition='Buono',
         address='Corso Vittorio Emanuele II', latitude=45.0625, longitude=7.6819,
         height='18', trunk_diameter_cm=62, crown_diameter_m=12, age='45',
         location='Filare stradale',
         tree_height_m=18, circonferenza_cm=195, branch_diam_cm=22, branch_length_m=8,
         branch_height_m=9, target_height_m=1.7,
         pericolo_rami='1', pericolo_tronco='0', pericolo_colletto='0', pericolo_zolla='0',
         bersaglio_chioma=6, bersaglio_ramo=5, moltiplicatore=1,
         comments='Esemplare in filare lungo corso trafficato.'),
    dict(custom_id='DEMO-002', species='Tilia cordata', condition='Discreto',
         address='Giardini Reali', latitude=45.0728, longitude=7.6874,
         height='14', trunk_diameter_cm=48, crown_diameter_m=9, age='35',
         location='Parco storico',
         tree_height_m=14, circonferenza_cm=151, branch_diam_cm=16, branch_length_m=6,
         branch_height_m=7, target_height_m=1.7,
         pericolo_rami='1', pericolo_tronco='1', pericolo_colletto='0', pericolo_zolla='0',
         bersaglio_chioma=4, bersaglio_ramo=3, moltiplicatore=1,
         comments='Lieve carie al colletto da monitorare.'),
    dict(custom_id='DEMO-003', species='Aesculus hippocastanum', condition='Scarso',
         address='Parco del Valentino', latitude=45.0556, longitude=7.6862,
         height='16', trunk_diameter_cm=55, crown_diameter_m=11, age='60',
         location='Parco urbano',
         tree_height_m=16, circonferenza_cm=173, branch_diam_cm=20, branch_length_m=7,
         branch_height_m=8, target_height_m=1.7,
         pericolo_rami='2', pericolo_tronco='2', pericolo_colletto='1', pericolo_zolla='1',
         bersaglio_chioma=5, bersaglio_ramo=4, moltiplicatore=1,
         comments='Corpi fruttiferi e defogliazione da cameraria.'),
    dict(custom_id='DEMO-004', species='Celtis australis', condition='Ottimo',
         address='Piazza Vittorio Veneto', latitude=45.0647, longitude=7.6942,
         height='12', trunk_diameter_cm=38, crown_diameter_m=8, age='25',
         location='Piazza', comments='Giovane esemplare in buona vigoria.'),
    dict(custom_id='DEMO-005', species='Quercus robur', condition='Buono',
         address='Parco della Pellerina', latitude=45.0889, longitude=7.6467,
         height='20', trunk_diameter_cm=78, crown_diameter_m=15, age='90',
         location='Grande parco', comments='Esemplare monumentale.'),
    dict(custom_id='DEMO-006', species='Cedrus deodara', condition='Discreto',
         address='Parco del Valentino', latitude=45.0533, longitude=7.6875,
         height='22', trunk_diameter_cm=70, crown_diameter_m=13, age='70',
         location='Parco urbano', comments='Conifera ornamentale, sofferenza estiva.'),
]

def _ensure_demo_account(username, password, role):
    """Create the demo account if missing; keep password/role/city in sync with config."""
    u = User.query.filter_by(username=username).first()
    if not u:
        u = User(username=username, email=None, role=role, city=DEMO_CITY_NAME)
        db.session.add(u)
    u.set_password(password)
    u.role = role
    u.city = DEMO_CITY_NAME
    db.session.commit()
    return u

def _seed_demo_trees(demo_user):
    """Delete all demo-owned trees (and their inspections) and re-create the fixed set."""
    tree_ids = [tid for (tid,) in
                db.session.query(Tree.id).filter(Tree.owner_id == demo_user.id).all()]
    if tree_ids:
        Inspection.query.filter(Inspection.tree_id.in_(tree_ids)).delete(synchronize_session=False)
        Tree.query.filter(Tree.owner_id == demo_user.id).delete(synchronize_session=False)
        db.session.commit()
    for d in DEMO_TREES:
        t = Tree(owner_id=demo_user.id, city=DEMO_CITY_NAME, **d)
        t.rischio = json.dumps(r) if (r := _calc_rischio(t)) else None
        db.session.add(t)
    db.session.commit()
    # Ispezione di "primo inserimento" per ogni albero demo: è il punto di
    # partenza che deve comparire nello storico.
    for t in Tree.query.filter(Tree.owner_id == demo_user.id).all():
        db.session.add(_make_inspection(t, demo_user.id, DEMO_USERNAME))
    db.session.commit()

def reset_demo():
    """Reset the whole demo to its fixed state: both accounts, their link, and the trees."""
    demo_user = _ensure_demo_account(DEMO_USERNAME, DEMO_PASSWORD, 'user')
    demo_city = _ensure_demo_account(DEMO_CITY_USERNAME, DEMO_CITY_PASSWORD, 'city')
    # demo_city manages only demo_user: clear any other links, then link demo_user.
    CityMembership.query.filter_by(city_user_id=demo_city.id).delete(synchronize_session=False)
    db.session.commit()
    db.session.add(CityMembership(city_user_id=demo_city.id, agronomer_id=demo_user.id))
    db.session.commit()
    _seed_demo_trees(demo_user)

def _demo_user_id():
    u = User.query.filter_by(username=DEMO_USERNAME).first()
    return u.id if u else None

def _is_demo_account(username):
    """Demo accounts are read-only for admin actions (creating users, linking agronomers)."""
    return username in (DEMO_USERNAME, DEMO_CITY_USERNAME)

def generate_token(user, expires_minutes=60*24):
    payload = {
        "user_id": user.id, "username": user.username,
        "role": user.role, "city": user.city,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=expires_minutes)
    }
    token = jwt.encode(payload, SECRET_KEY, algorithm="HS256")
    return token.decode('utf-8') if isinstance(token, bytes) else token

def auth_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        auth = request.headers.get('Authorization', None)
        if not auth:
            return jsonify({'message': 'Missing Authorization header'}), 401
        token = auth.split(" ", 1)[1] if auth.startswith("Bearer ") else auth
        try:
            data = jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
        except jwt.ExpiredSignatureError:
            return jsonify({'message': 'Token expired'}), 401
        except Exception as e:
            return jsonify({'message': 'Invalid token', 'error': str(e)}), 401
        request.user = data
        return f(*args, **kwargs)
    return wrapper

def role_required(*allowed_roles):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if not hasattr(request, 'user'):
                return jsonify({'message': 'Unauthorized'}), 401
            if request.user.get('role') not in allowed_roles:
                return jsonify({'message': 'Forbidden - insufficient role'}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

# -----------------------
# Auth endpoints
# -----------------------
@app.route('/login', methods=['POST'])
def login():
    data = request.json or {}
    username, password = data.get('username'), data.get('password')
    if not username or not password:
        return jsonify({'message': 'username and password required'}), 400
    user = User.query.filter_by(username=username).first()
    if not user or not user.check_password(password):
        return jsonify({'message': 'Invalid credentials'}), 401
    # Demo accounts: reset the shared sandbox to the fixed demo state on every login.
    if user.username in (DEMO_USERNAME, DEMO_CITY_USERNAME):
        reset_demo()
    token = generate_token(user)
    return jsonify({'token': token,
                    'user': {'id': user.id, 'username': user.username,
                             'role': user.role, 'city': user.city}})

@app.route('/me', methods=['GET'])
@auth_required
def me():
    return jsonify({'user_id': request.user.get('user_id'),
                    'username': request.user.get('username'),
                    'role': request.user.get('role'),
                    'city': request.user.get('city')})

@app.route('/register', methods=['POST'])
def register():
    if not REGISTRATION_ENABLED:
        return jsonify({'message': 'Registrazione non disponibile'}), 403
    data = request.json or {}
    username = data.get('username', '').strip()
    email    = data.get('email', '').strip()
    password = data.get('password', '')
    if not username or not password:
        return jsonify({'message': 'Username e password obbligatori'}), 400
    if len(password) < 6:
        return jsonify({'message': 'Password troppo corta (minimo 6 caratteri)'}), 400
    if User.query.filter_by(username=username).first():
        return jsonify({'message': 'Username già in uso'}), 400
    if email and User.query.filter_by(email=email).first():
        return jsonify({'message': 'Email già registrata'}), 400
    user = User(username=username, email=email or None, role='user', city=None)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    token = generate_token(user)
    return jsonify({'token': token,
                    'user': {'id': user.id, 'username': user.username,
                             'role': user.role, 'city': user.city}}), 201

@app.route('/forgot-password', methods=['POST'])
def forgot_password():
    data  = request.json or {}
    email = data.get('email', '').strip()
    user  = User.query.filter_by(email=email).first() if email else None
    if user:
        token = secrets.token_urlsafe(32)
        user.password_reset_token   = token
        user.password_reset_expires = datetime.now(timezone.utc) + timedelta(hours=2)
        db.session.commit()
        reset_url = f"{APP_BASE_URL}/?token={token}"
        if SMTP_USER and SMTP_PASSWORD:
            try:
                html_body = f"""
<div style="font-family:sans-serif;max-width:480px;margin:0 auto;padding:32px 24px;background:#f9f9f7;border-radius:12px;">
  <h2 style="color:#2d6a4f;margin-top:0;">Reimposta la tua password</h2>
  <p style="color:#444;">Hai richiesto il reset della password per il tuo account <strong>Silvae Pro</strong>.</p>
  <p style="color:#444;">Clicca il pulsante qui sotto per scegliere una nuova password. Il link è valido per <strong>2 ore</strong>.</p>
  <a href="{reset_url}"
     style="display:inline-block;margin:20px 0;padding:12px 28px;background:#2d6a4f;color:#fff;text-decoration:none;border-radius:8px;font-weight:bold;">
    Reimposta password
  </a>
  <p style="font-size:12px;color:#888;">Se non hai richiesto il reset, ignora questa email. La tua password rimarrà invariata.</p>
  <hr style="border:none;border-top:1px solid #ddd;margin:24px 0;">
  <p style="font-size:11px;color:#aaa;">Silvae Pro &nbsp;·&nbsp; Sistema di Gestione Alberi</p>
</div>
"""
                msg = MIMEMultipart('alternative')
                msg['Subject'] = 'Silvae Pro – Reimposta la tua password'
                msg['From']    = SMTP_USER
                msg['To']      = user.email
                msg.attach(MIMEText(f"Reimposta la password: {reset_url}", 'plain'))
                msg.attach(MIMEText(html_body, 'html'))
                with smtplib.SMTP('smtp.gmail.com', 587) as smtp:
                    smtp.ehlo()
                    smtp.starttls()
                    smtp.login(SMTP_USER, SMTP_PASSWORD)
                    smtp.sendmail(SMTP_USER, user.email, msg.as_string())
            except Exception as e:
                print(f"[PASSWORD RESET] Errore invio email: {e}")
        else:
            print(f"[PASSWORD RESET] SMTP non configurato. Link: {reset_url}")
    # Always return the same message to avoid user enumeration
    return jsonify({'message': 'Se l\'email è registrata riceverai un link per il reset.'}), 200

@app.route('/reset-password', methods=['POST'])
def reset_password():
    data         = request.json or {}
    token        = data.get('token', '').strip()
    new_password = data.get('password', '')
    if not token or not new_password:
        return jsonify({'message': 'Token e nuova password obbligatori'}), 400
    if len(new_password) < 6:
        return jsonify({'message': 'Password troppo corta (minimo 6 caratteri)'}), 400
    user = User.query.filter_by(password_reset_token=token).first()
    if not user or not user.password_reset_expires:
        return jsonify({'message': 'Token non valido'}), 400
    if user.password_reset_expires < datetime.now(timezone.utc):
        return jsonify({'message': 'Token scaduto'}), 400
    user.set_password(new_password)
    user.password_reset_token   = None
    user.password_reset_expires = None
    db.session.commit()
    return jsonify({'message': 'Password aggiornata con successo'}), 200

# -----------------------
# User management
# -----------------------
@app.route('/add_user', methods=['POST'])
@auth_required
def add_user():
    creator = request.user
    if _is_demo_account(creator.get('username')):
        return jsonify({'message': 'Azione non disponibile nell\'account demo'}), 403
    data = request.json or {}
    username, password, role, city = (data.get(k) for k in ('username','password','role','city'))
    if not username or not password or not role:
        return jsonify({'message': 'username, password and role are required'}), 400
    if creator['role'] == 'user':
        return jsonify({'message': 'Users cannot create new users'}), 403
    if creator['role'] == 'city':
        if role != 'user':
            return jsonify({'message': 'City users can only create role "user"'}), 403
        city = creator.get('city')
    if User.query.filter_by(username=username).first():
        return jsonify({'message': 'Username already exists'}), 400
    new_user = User(username=username, role=role, city=city)
    new_user.set_password(password)
    db.session.add(new_user)
    db.session.commit()
    return jsonify({'message': 'User created',
                    'user': {'id': new_user.id, 'username': username, 'role': role, 'city': city}}), 201

@app.route('/users', methods=['GET'])
@auth_required
def list_users():
    role, user_city, user_id = (request.user.get(k) for k in ('role','city','user_id'))
    if role == 'superuser':   users = User.query.all()
    elif role == 'city':      users = User.query.filter_by(city=user_city).all()
    else:                     users = User.query.filter_by(id=user_id).all()
    return jsonify([{'id': u.id, 'username': u.username, 'role': u.role, 'city': u.city} for u in users])

# -----------------------
# City ↔ Agronomer membership
# -----------------------
@app.route('/city/agronomers', methods=['GET'])
@auth_required
@role_required('city', 'superuser')
def list_city_agronomers():
    user_id = request.user.get('user_id')
    role    = request.user.get('role')
    city_user_id = int(request.args.get('city_user_id', user_id)) if role == 'superuser' else user_id
    rows = CityMembership.query.filter_by(city_user_id=city_user_id).all()
    result = []
    for m in rows:
        a = db.session.get(User, m.agronomer_id)
        if a:
            result.append({'membership_id': m.id, 'agronomer_id': a.id,
                           'username': a.username, 'email': a.email or ''})
    return jsonify(result)

@app.route('/city/agronomers', methods=['POST'])
@auth_required
@role_required('city', 'superuser')
def add_city_agronomer():
    if _is_demo_account(request.user.get('username')):
        return jsonify({'message': 'Azione non disponibile nell\'account demo'}), 403
    user_id = request.user.get('user_id')
    role    = request.user.get('role')
    data    = request.json or {}
    username = data.get('username', '').strip()
    if not username:
        return jsonify({'message': 'username obbligatorio'}), 400
    agronomer = User.query.filter_by(username=username, role='user').first()
    if not agronomer:
        return jsonify({'message': f'Agronomo "{username}" non trovato'}), 404
    city_user_id = user_id if role == 'city' else int(data.get('city_user_id', user_id))
    if CityMembership.query.filter_by(city_user_id=city_user_id, agronomer_id=agronomer.id).first():
        return jsonify({'message': 'Agronomo già collegato a questo comune'}), 409
    m = CityMembership(city_user_id=city_user_id, agronomer_id=agronomer.id)
    db.session.add(m)
    db.session.commit()
    return jsonify({'membership_id': m.id, 'agronomer_id': agronomer.id,
                    'username': agronomer.username, 'email': agronomer.email or ''}), 201

@app.route('/city/agronomers/<int:membership_id>', methods=['DELETE'])
@auth_required
@role_required('city', 'superuser')
def remove_city_agronomer(membership_id):
    if _is_demo_account(request.user.get('username')):
        return jsonify({'message': 'Azione non disponibile nell\'account demo'}), 403
    user_id = request.user.get('user_id')
    role    = request.user.get('role')
    m = db.session.get(CityMembership, membership_id)
    if not m:
        return jsonify({'message': 'Collegamento non trovato'}), 404
    if role == 'city' and m.city_user_id != user_id:
        return jsonify({'message': 'Forbidden'}), 403
    db.session.delete(m)
    db.session.commit()
    return jsonify({'message': 'Agronomo rimosso dal comune'}), 200

# -----------------------
# City management
# -----------------------
@app.route('/admin/cities', methods=['POST'])
@auth_required
@role_required('superuser')
def add_city():
    name = (request.json or {}).get('name')
    if not name: return jsonify({'message': 'City name required'}), 400
    if City.query.filter_by(name=name).first():
        return jsonify({'message': 'City already exists'}), 400
    c = City(name=name)
    db.session.add(c); db.session.commit()
    return jsonify({'message': 'City created', 'city': {'id': c.id, 'name': c.name}}), 201

@app.route('/cities', methods=['GET'])
def get_cities():
    from_table = {c.name for c in City.query.all()}
    from_trees = {row[0] for row in db.session.query(Tree.city).distinct() if row[0]}
    return jsonify(sorted(from_table | from_trees))

@app.route('/comuni/search', methods=['GET'])
def search_comuni():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])
    results = (
        Comune.query
        .filter(Comune.nome.ilike(f'%{q}%'))
        .order_by(
            case((Comune.nome.ilike(f'{q}%'), 0), else_=1),
            Comune.nome
        )
        .limit(20)
        .all()
    )
    return jsonify([{
        'nome': c.nome,
        'provincia': c.provincia,
        'sigla': c.sigla,
        'regione': c.regione,
    } for c in results])

# -----------------------
# Dropdowns endpoint
# -----------------------
@app.route('/dropdowns', methods=['GET'])
def get_dropdowns():
    return jsonify({
        'species': [{'code': k, 'name': v} for k, v in species_lookup.items()],
        'dimora': dd.dimora,
        'stadio_sviluppo': dd.stadio_sviluppo,
        'posizione_sociale': dd.posizione_sociale,
        'localizzazione': dd.localizzazione,
        'vincoli': dd.vincoli,
        'giudizio_severita': dd.giudizio_severita,
        'pericolo_ord': {str(k): v for k, v in dd.pericolo_ord.items()},
        'bersaglio_tipi': dd.BERSAGLIO_TIPI,
        'bersaglio_proprieta_values': dd.BERSAGLIO_PROPRIETA_VALUES,
        'bersaglio_occupazione_values': dd.BERSAGLIO_OCCUPAZIONE_VALUES,
        'condizione_salute_ecologica': [desc for desc, _ in dd.CONDIZIONE_SALUTE_ECOLOGICA],
        'bersaglio_range': list(range(1, 8)),
        'diag_zolla': dd.zolla_radicale,
        'diag_colletto': dd.caratteri_colletto,
        'diag_fusto': dd.caratteri_fusto,
        'diag_castello': dd.caratteri_castello,
        'diag_ramificazione': dd.caratteri_ramificazione,
        'diag_chioma': dd.caratteri_chioma,
        'monitoraggio': lt_monitoraggio,
        'urgenza': lt_urgenza,
        'conflitti': lt_conflitti,
        'agenti_carie': lt_agenti_carie,
        'altri_patogeni': lt_altri_patogeni,
        'prescrizioni_val': lt_prescrizioni_val,
        'prescrizioni_mit': lt_prescrizioni_mit,
        'prescrizioni_col': lt_prescrizioni_col,
    })

# -----------------------
# Risk calculation
# -----------------------
@app.route('/calculate_risk', methods=['POST'])
@auth_required
def calculate_risk():
    data = request.json or {}
    def f(k):
        v = data.get(k)
        try: return float(v) if v not in (None,'') else None
        except: return None
    def i(k):
        v = data.get(k)
        try: return int(v) if v not in (None,'') else None
        except: return None

    required = {
        'tree_height_m': f('tree_height_m'), 'circonferenza_cm': f('circonferenza_cm'),
        'branch_diam_cm': f('branch_diam_cm'), 'branch_length_m': f('branch_length_m'),
        'branch_height_m': f('branch_height_m'), 'target_height_m': f('target_height_m'),
        'pericolo_rami': i('pericolo_rami'), 'pericolo_tronco': i('pericolo_tronco'),
        'pericolo_colletto': i('pericolo_colletto'), 'pericolo_zolla': i('pericolo_zolla'),
        'bersaglio_chioma': i('bersaglio_chioma'), 'bersaglio_ramo': i('bersaglio_ramo'),
    }
    # Resolve bersaglio class from tipo+value (proprietà/occupazione)
    # or from tipo+flow (pedoni/traffico).
    crown_diam_m = f('crown_diameter_m') or 0
    tree_h_m     = required['tree_height_m'] or 0
    branch_l_m   = required['branch_length_m'] or 0
    zone_chioma  = (crown_diam_m + tree_h_m) / 2.0
    zone_ramo    = branch_l_m * 1.25
    for tipo_key, value_key, flow_key, class_key, zone_w in [
        ('bersaglio_chioma_tipo', 'bersaglio_chioma_value', 'bersaglio_chioma_flow',
         'bersaglio_chioma', zone_chioma),
        ('bersaglio_ramo_tipo',   'bersaglio_ramo_value',   'bersaglio_ramo_flow',
         'bersaglio_ramo',   zone_ramo),
    ]:
        tipo  = data.get(tipo_key)
        value = data.get(value_key)
        flow  = f(flow_key)
        if tipo and value:
            computed = bersaglio_value_to_class(tipo, value)
            if computed is not None:
                required[class_key] = computed
        elif tipo and flow is not None and zone_w > 0:
            computed = bersaglio_flow_to_class(tipo, flow, zone_w)
            if computed is not None:
                required[class_key] = computed
    # I gradi di pericolo si valutano per-componente, come nella scheda: un
    # pericolo mancante diventa 0 = "non determinato" (SOSPESO), non blocca il
    # calcolo. Si rifiuta solo se mancano TUTTI e quattro.
    PERICOLI = ('pericolo_rami', 'pericolo_tronco', 'pericolo_colletto', 'pericolo_zolla')
    if all(required[k] is None for k in PERICOLI):
        return jsonify({'message': 'Per calcolare il rischio inserisci almeno un '
                                   'grado di pericolo (rami, tronco, colletto o zolla).'}), 400
    for k in PERICOLI:
        if required[k] is None:
            required[k] = 0

    # Gli altri campi (dimensioni e bersagli) restano obbligatori.
    missing = [k for k, v in required.items() if v is None]
    if missing:
        return jsonify({'message': f'Missing: {", ".join(missing)}'}), 400

    result = assess_tree(
        tree_height_m=required['tree_height_m'],
        circumference_cm=required['circonferenza_cm'],
        crown_diam_m=crown_diam_m,
        branch_diam_cm=required['branch_diam_cm'],
        branch_length_m=required['branch_length_m'],
        branch_height_m=required['branch_height_m'],
        target_height_m=required['target_height_m'],
        pericolo_rami=required['pericolo_rami'],
        pericolo_tronco=required['pericolo_tronco'],
        pericolo_colletto=required['pericolo_colletto'],
        pericolo_zolla=required['pericolo_zolla'],
        bersaglio_chioma=required['bersaglio_chioma'],
        bersaglio_ramo=required['bersaglio_ramo'],
        moltiplicatore=i('moltiplicatore') or 1,
        post_tree_height_m=f('post_tree_height_m'),
        post_circumference_cm=f('post_circonferenza_cm'),
        post_branch_diam_cm=f('post_branch_diam_cm'),
        post_branch_length_m=f('post_branch_length_m'),
        post_branch_height_m=f('post_branch_height_m'),
        post_target_height_m=f('post_target_height_m'),
    )
    return jsonify(result)

# -----------------------
# City visibility helpers
# -----------------------
def city_tree_filter(query, city_user_id, user_city):
    linked = db.session.query(CityMembership.agronomer_id).filter_by(city_user_id=city_user_id)
    return query.filter(Tree.owner_id.in_(linked), Tree.city == user_city)

def city_can_access_tree(tree, city_user_id, user_city):
    if tree.owner_id is None or tree.city != user_city:
        return False
    return CityMembership.query.filter_by(
        city_user_id=city_user_id, agronomer_id=tree.owner_id).first() is not None

# -----------------------
# Tree endpoints
# -----------------------
@app.route('/trees', methods=['GET'])
@auth_required
def get_trees():
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    query = Tree.query
    if role == 'user':   query = query.filter(Tree.owner_id == user_id)
    elif role == 'city': query = city_tree_filter(query, user_id, user_city)
    elif role == 'superuser':
        did = _demo_user_id()
        if did and did != user_id: query = query.filter(Tree.owner_id != did)
    city_q = request.args.get('city')
    addr_q = request.args.get('address')
    if city_q:  query = query.filter(Tree.city.ilike(f"%{city_q}%"))
    if addr_q:  query = query.filter(Tree.address.ilike(f"%{addr_q}%"))
    return jsonify([tree_to_dict(t) for t in query.order_by(Tree.id).all()])

@app.route('/tree/<int:tree_id>', methods=['GET'])
@auth_required
def get_tree_by_id(tree_id):
    tree = db.session.get(Tree, tree_id)
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    if role == 'user' and tree.owner_id != user_id:          return jsonify({'message': 'Forbidden'}), 403
    if role == 'city' and not city_can_access_tree(tree, user_id, user_city): return jsonify({'message': 'Forbidden'}), 403
    return jsonify(tree_to_dict(tree))

@app.route('/tree/<int:tree_id>', methods=['PATCH'])
@auth_required
def update_tree(tree_id):
    tree = db.session.get(Tree, tree_id)
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    if role == 'user' and tree.owner_id != user_id:          return jsonify({'message': 'Forbidden'}), 403
    if role == 'city' and not city_can_access_tree(tree, user_id, user_city): return jsonify({'message': 'Forbidden'}), 403

    data = request.json or {}
    apply_tree_fields(tree, data)

    if 'next_check' in data and data['next_check']:
        try: tree.next_check = datetime.strptime(data['next_check'], "%Y-%m-%d")
        except ValueError: return jsonify({'message': 'Invalid date for next_check'}), 400

    if data.get('latitude') and data.get('longitude'):
        try:
            tree.latitude  = float(data['latitude'])
            tree.longitude = float(data['longitude'])
        except (ValueError, TypeError):
            return jsonify({'message': 'Invalid lat/lon'}), 400

    if tree.cpc:
        tree.condition = tree.cpc

    tree.rischio = json.dumps(r) if (r := _calc_rischio(tree)) else None
    db.session.commit()

    if {'condition','comments','actions'} & set(data.keys()):
        db.session.add(_make_inspection(tree, request.user.get('user_id'), request.user.get('username')))
        db.session.commit()
    return jsonify({'message': f'Tree {tree_id} updated successfully!'}), 200

@app.route('/tree/<int:tree_id>', methods=['DELETE'])
@auth_required
def delete_tree(tree_id):
    tree = db.session.get(Tree, tree_id)
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    if role == 'user' and tree.owner_id != user_id:          return jsonify({'message': 'Forbidden'}), 403
    if role == 'city' and not city_can_access_tree(tree, user_id, user_city): return jsonify({'message': 'Forbidden'}), 403
    db.session.delete(tree); db.session.commit()
    return jsonify({'message': f'Tree {tree_id} deleted successfully!'}), 200

@app.route('/trees/bulk', methods=['DELETE'])
@auth_required
def delete_trees_bulk():
    user_id   = request.user.get('user_id')
    role      = request.user.get('role')
    user_city = request.user.get('city')
    ids = (request.json or {}).get('ids', [])
    if not ids:
        return jsonify({'message': 'Nessun ID fornito'}), 400
    deleted = 0
    for tree_id in ids:
        tree = db.session.get(Tree, int(tree_id))
        if not tree: continue
        if role == 'user' and tree.owner_id != user_id: continue
        if role == 'city' and not city_can_access_tree(tree, user_id, user_city): continue
        db.session.delete(tree)
        deleted += 1
    db.session.commit()
    return jsonify({'deleted': deleted})

@app.route('/add_tree', methods=['POST'])
@auth_required
def add_tree():
    data = request.json or {}
    user_id = request.user.get('user_id')
    role    = request.user.get('role')
    user_city = request.user.get('city')

    if not data.get('custom_id') or not data.get('species'):
        return jsonify({'message': 'custom_id and species are required'}), 400

    if not data.get('city'):
        if role in ('user','city') and user_city: data['city'] = user_city
        else: return jsonify({'message': 'city is required'}), 400

    if not data.get('latitude') or not data.get('longitude'):
        full_address = (data.get('address') or '') + ' ' + data.get('city','')
        loc = geolocator.geocode(full_address)
        if loc: data['latitude'], data['longitude'] = loc.latitude, loc.longitude
        else: return jsonify({'error': 'Invalid address, coordinates not found'}), 400

    try:
        next_check_date = datetime.strptime(data['next_check'], "%Y-%m-%d") if data.get('next_check') else None
    except ValueError:
        return jsonify({'message': 'Invalid date for next_check'}), 400

    new_tree = Tree(
        custom_id=data['custom_id'],
        latitude=float(data['latitude']), longitude=float(data['longitude']),
        address=data.get('address',''), city=data['city'],
        species=data['species'], condition=data.get('condition','—'),
        comments=data.get('comments',''), actions=data.get('actions',''),
        height=data.get('height',''), trunk_diameter_cm=data.get('trunk_diameter_cm'),
        crown_diameter_m=data.get('crown_diameter_m'), age=data.get('age',''),
        location=data.get('location',''), cpc=data.get('cpc',''),
        next_check=next_check_date,
        owner_id=user_id
    )
    apply_tree_fields(new_tree, data)
    if new_tree.cpc:
        new_tree.condition = new_tree.cpc
    new_tree.rischio = json.dumps(r) if (r := _calc_rischio(new_tree)) else None
    db.session.add(new_tree)
    try:
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'message': 'Errore durante il salvataggio. Riprova.'}), 409

    db.session.add(_make_inspection(new_tree, user_id, request.user.get('username')))
    db.session.commit()
    return jsonify({'message': 'Tree added successfully!', 'tree_id': new_tree.id}), 201

@app.route('/tree/custom/<string:custom_id>', methods=['GET'])
@auth_required
def get_tree_by_custom_id(custom_id):
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    query = Tree.query.filter_by(custom_id=custom_id)
    if role == 'user':   query = query.filter(Tree.owner_id == user_id)
    elif role == 'city': query = city_tree_filter(query, user_id, user_city)
    else:
        city_param = request.args.get('city')
        if city_param: query = query.filter(Tree.city == city_param)
    tree = query.first()
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    return jsonify(tree_to_dict(tree))

@app.route('/tree/<int:tree_id>/inspections', methods=['GET'])
@auth_required
def get_inspections(tree_id):
    tree = db.session.get(Tree, tree_id)
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    if role == 'user' and tree.owner_id != user_id:          return jsonify({'message': 'Forbidden'}), 403
    if role == 'city' and not city_can_access_tree(tree, user_id, user_city): return jsonify({'message': 'Forbidden'}), 403
    rows = (Inspection.query.filter_by(tree_id=tree_id)
            .order_by(Inspection.date.desc(), Inspection.created_at.desc()).all())
    return jsonify([{
        'id': i.id, 'date': i.date.strftime('%Y-%m-%d'),
        'condition': i.condition, 'comments': i.comments, 'actions': i.actions,
        'inspector_name': i.inspector_name,
        'created_at': i.created_at.strftime('%Y-%m-%d %H:%M') if i.created_at else None,
        'rischio': json.loads(i.rischio) if i.rischio else None,
        'snapshot': json.loads(i.snapshot) if i.snapshot else None,
    } for i in rows])

@app.route('/tree/<int:tree_id>/inspections', methods=['POST'])
@auth_required
def add_inspection(tree_id):
    tree = db.session.get(Tree, tree_id)
    if not tree: return jsonify({'message': 'Tree not found'}), 404
    role, user_id, user_city = (request.user.get(k) for k in ('role','user_id','city'))
    if role == 'user' and tree.owner_id != user_id:          return jsonify({'message': 'Forbidden'}), 403
    if role == 'city' and not city_can_access_tree(tree, user_id, user_city): return jsonify({'message': 'Forbidden'}), 403
    data = request.json or {}
    date_str = data.get('date') or datetime.now(timezone.utc).strftime('%Y-%m-%d')
    try: insp_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError: return jsonify({'message': 'Invalid date format'}), 400
    condition = data.get('condition') or tree.condition
    rischio_val = _calc_rischio(tree)
    snap = tree_to_dict(tree)
    snap['condition'] = condition

    insp = Inspection(
        tree_id=tree_id, date=insp_date,
        condition=condition,
        comments=data.get('comments',''), actions=data.get('actions',''),
        inspector_id=user_id, inspector_name=request.user.get('username'),
        snapshot=json.dumps(snap),
        rischio=json.dumps(rischio_val) if rischio_val else None,
    )
    db.session.add(insp); db.session.commit()
    return jsonify({'message': 'Inspection logged', 'id': insp.id}), 201

@app.route('/test_geocode', methods=['POST'])
def test_geocode():
    address = request.json.get('address')
    loc = geolocator.geocode(address)
    if loc: return jsonify({'latitude': loc.latitude, 'longitude': loc.longitude})
    return jsonify({'error': 'Address not found'}), 404

@app.route('/reverse_geocode', methods=['POST'])
@auth_required
def reverse_geocode():
    data = request.json or {}
    lat, lon = data.get('latitude'), data.get('longitude')
    if lat is None or lon is None:
        return jsonify({'message': 'latitude and longitude required'}), 400
    try:
        loc = geolocator.reverse(f"{lat}, {lon}", language='en')
        if loc:
            addr = loc.raw.get('address', {})
            city = (addr.get('city') or addr.get('town') or
                    addr.get('village') or addr.get('municipality') or '')
            return jsonify({'address': loc.address, 'city': city})
    except Exception:
        pass
    return jsonify({'address': '', 'city': ''})

# --- Excel export: helper e definizione delle colonne (raggruppate) -----------
# Definiti a livello di modulo così da poter essere condivisi tra l'endpoint di
# export e quello che elenca le colonne (per la scelta dei campi da esportare).
def _xl_jlist(val):
    if not val: return ''
    try:
        lst = json.loads(val) if isinstance(val, str) else val
        return ', '.join(str(x) for x in lst) if isinstance(lst, list) else str(lst)
    except: return ''

def _xl_diag(val):
    if not val: return ''
    try:
        lst = json.loads(val) if isinstance(val, str) else val
        return '; '.join(
            f"{r.get('caratt','')}: {r.get('giudizio','')}"
            for r in lst if isinstance(r, dict) and r.get('caratt')
        )
    except: return ''

def _xl_rtop(t, phase, field):
    if not t.rischio: return ''
    try: return json.loads(t.rischio).get(phase, {}).get(field, '')
    except: return ''

def _xl_rpart(t, phase, part, field):
    if not t.rischio: return ''
    try: return json.loads(t.rischio).get(phase, {}).get(part, {}).get(field, '')
    except: return ''


def _excel_groups():
    """Colonne dell'export Excel: (group_label, color_hex, [(col_header, getter), ...]).
    L'intestazione di colonna funge anche da chiave stabile per l'esclusione dei campi."""
    _jlist, _diag, _rtop, _rpart = _xl_jlist, _xl_diag, _xl_rtop, _xl_rpart
    return [
        ('Identificazione', 'BDD7EE', [
            ('ID',                  lambda t: t.id),
            ('ID Albero',           lambda t: t.custom_id),
            ('Comune',              lambda t: t.city),
            ('Indirizzo',           lambda t: t.address or ''),
            ('Latitudine',          lambda t: t.latitude),
            ('Longitudine',         lambda t: t.longitude),
            ('Specie',              lambda t: t.species),
            ('Condizione',          lambda t: t.condition),
            ('Età',                 lambda t: t.age or ''),
            ('CPC',                 lambda t: t.cpc or ''),
            ('Posizione',           lambda t: t.location or ''),
            ('Dimora',              lambda t: t.dimora or ''),
            ('Stadio sviluppo',     lambda t: t.stadio_sviluppo or ''),
            ('Posizione sociale',   lambda t: t.posizione_sociale or ''),
            ('Localizzazione',      lambda t: t.localizzazione or ''),
            ('Vincoli',             lambda t: t.vincoli or ''),
            ('Prossima ispezione',  lambda t: t.next_check.strftime('%Y-%m-%d') if t.next_check else ''),
            ('Urgenza',             lambda t: t.urgenza or ''),
            ('Monitoraggio',        lambda t: t.monitoraggio or ''),
            ('Note',                lambda t: t.comments or ''),
            ('Azioni',              lambda t: t.actions or ''),
        ]),
        ('Misure', 'C6EFCE', [
            ('Altezza (m)',         lambda t: t.tree_height_m),
            ('Altezza generica',    lambda t: t.height or ''),
            ('Circonferenza (cm)',  lambda t: t.circonferenza_cm),
            ('Ø Tronco (cm)',       lambda t: t.trunk_diameter_cm),
            ('Ø Chioma (m)',        lambda t: t.crown_diameter_m),
            ('Ø Ramo (cm)',         lambda t: t.branch_diam_cm),
            ('Lung. ramo (m)',      lambda t: t.branch_length_m),
            ('Alt. ramo (m)',       lambda t: t.branch_height_m),
            ('Alt. target (m)',     lambda t: t.target_height_m),
        ]),
        ('Post-intervento', 'FFF2CC', [
            ('Altezza post (m)',        lambda t: t.post_tree_height_m),
            ('Circonferenza post (cm)', lambda t: t.post_circonferenza_cm),
            ('Ø Ramo post (cm)',        lambda t: t.post_branch_diam_cm),
            ('Lung. ramo post (m)',     lambda t: t.post_branch_length_m),
            ('Alt. ramo post (m)',      lambda t: t.post_branch_height_m),
            ('Alt. target post (m)',    lambda t: t.post_target_height_m),
        ]),
        ('Valutazione rischio', 'FCE4D6', [
            ('Pericolo rami',               lambda t: t.pericolo_rami or ''),
            ('Pericolo tronco',             lambda t: t.pericolo_tronco or ''),
            ('Pericolo colletto',           lambda t: t.pericolo_colletto or ''),
            ('Pericolo zolla',              lambda t: t.pericolo_zolla or ''),
            ('Tipo bersaglio chioma',        lambda t: t.bersaglio_chioma_tipo or ''),
            ('Descrizione bersaglio chioma',lambda t: t.bersaglio_chioma_value or ''),
            ('Flusso bersaglio chioma',     lambda t: t.bersaglio_chioma_flow),
            ('Classe bersaglio chioma',     lambda t: t.bersaglio_chioma),
            ('Tipo bersaglio ramo',         lambda t: t.bersaglio_ramo_tipo or ''),
            ('Descrizione bersaglio ramo',  lambda t: t.bersaglio_ramo_value or ''),
            ('Flusso bersaglio ramo',       lambda t: t.bersaglio_ramo_flow),
            ('Classe bersaglio ramo',       lambda t: t.bersaglio_ramo),
            ('Moltiplicatore (n bersagli)', lambda t: t.moltiplicatore or 1),
            ('Imp. chioma att. (kgm/s)',    lambda t: _rtop(t, 'attuale', 'crown_momentum_kgms')),
            ('Classe imp. chioma att.',     lambda t: _rtop(t, 'attuale', 'crown_impulso_class')),
            ('Imp. ramo att. (kgm/s)',      lambda t: _rtop(t, 'attuale', 'branch_momentum_kgms')),
            ('Classe imp. ramo att.',       lambda t: _rtop(t, 'attuale', 'branch_impulso_class')),
            ('Rischio rami 1bers (att.)',   lambda t: _rpart(t, 'attuale', 'rami',     'risk_ratio_1bers')),
            ('Rischio rami (att.)',         lambda t: _rpart(t, 'attuale', 'rami',     'risk_description')),
            ('Rischio tronco 1bers (att.)',lambda t: _rpart(t, 'attuale', 'tronco',   'risk_ratio_1bers')),
            ('Rischio tronco (att.)',       lambda t: _rpart(t, 'attuale', 'tronco',   'risk_description')),
            ('Rischio colletto 1bers (att.)',lambda t: _rpart(t, 'attuale', 'colletto','risk_ratio_1bers')),
            ('Rischio colletto (att.)',     lambda t: _rpart(t, 'attuale', 'colletto', 'risk_description')),
            ('Rischio zolla 1bers (att.)', lambda t: _rpart(t, 'attuale', 'zolla',    'risk_ratio_1bers')),
            ('Rischio zolla (att.)',        lambda t: _rpart(t, 'attuale', 'zolla',    'risk_description')),
            ('Imp. chioma res. (kgm/s)',    lambda t: _rtop(t, 'residuo', 'crown_momentum_kgms')),
            ('Classe imp. chioma res.',     lambda t: _rtop(t, 'residuo', 'crown_impulso_class')),
            ('Imp. ramo res. (kgm/s)',      lambda t: _rtop(t, 'residuo', 'branch_momentum_kgms')),
            ('Classe imp. ramo res.',       lambda t: _rtop(t, 'residuo', 'branch_impulso_class')),
            ('Rischio rami 1bers (res.)',   lambda t: _rpart(t, 'residuo', 'rami',     'risk_ratio_1bers')),
            ('Rischio rami (res.)',         lambda t: _rpart(t, 'residuo', 'rami',     'risk_description')),
            ('Rischio tronco 1bers (res.)',lambda t: _rpart(t, 'residuo', 'tronco',   'risk_ratio_1bers')),
            ('Rischio tronco (res.)',       lambda t: _rpart(t, 'residuo', 'tronco',   'risk_description')),
            ('Rischio colletto 1bers (res.)',lambda t: _rpart(t, 'residuo', 'colletto','risk_ratio_1bers')),
            ('Rischio colletto (res.)',     lambda t: _rpart(t, 'residuo', 'colletto', 'risk_description')),
            ('Rischio zolla 1bers (res.)', lambda t: _rpart(t, 'residuo', 'zolla',    'risk_ratio_1bers')),
            ('Rischio zolla (res.)',        lambda t: _rpart(t, 'residuo', 'zolla',    'risk_description')),
        ]),
        ('Diagnosi', 'E2EFDA', [
            ('Diagnosi zolla',          lambda t: _diag(t.diag_zolla)),
            ('Diagnosi colletto',       lambda t: _diag(t.diag_colletto)),
            ('Diagnosi fusto',          lambda t: _diag(t.diag_fusto)),
            ('Diagnosi castello',       lambda t: _diag(t.diag_castello)),
            ('Diagnosi ramificazione',  lambda t: _diag(t.diag_ramificazione)),
            ('Diagnosi chioma',         lambda t: _diag(t.diag_chioma)),
        ]),
        ('Prescrizioni e patologie', 'E8DAEF', [
            ('Conflitti',                   lambda t: _jlist(t.conflitti_list)),
            ('Agenti carie',                lambda t: _jlist(t.agenti_carie)),
            ('Altri patogeni',              lambda t: _jlist(t.altri_patogeni)),
            ('Prescrizioni valutazione',    lambda t: _jlist(t.prescrizioni_val)),
            ('Prescrizioni mitigazione',    lambda t: _jlist(t.prescrizioni_mit)),
            ('Prescrizioni collaterale',    lambda t: _jlist(t.prescrizioni_col)),
        ]),
    ]


def _export_exclude_set():
    """Insieme dei campi da NON esportare, dal parametro `exclude` (array JSON di chiavi).
    La chiave è l'intestazione colonna (Excel) o il nome colonna di default (GPKG)."""
    raw = request.args.get('exclude')
    if not raw:
        return set()
    try:
        data = json.loads(raw)
        if isinstance(data, list):
            return {str(x) for x in data}
    except Exception:
        pass
    return set()


@app.route('/export/excel/columns', methods=['GET'])
@auth_required
def export_excel_columns():
    """Gruppi e intestazioni delle colonne dell'export Excel, per la scelta dei campi."""
    return jsonify([
        {'group': label, 'color': color, 'columns': [h for h, _ in cols]}
        for label, color, cols in _excel_groups()
    ])


@app.route('/export/excel', methods=['GET'])
@auth_required
def export_excel():
    trees   = _scoped_tree_query().all()
    exclude = _export_exclude_set()

    # Flatten columns and track group spans for merged header row,
    # skipping excluded columns and dropping any group left empty.
    all_cols   = []  # [(header, getter, color)]
    group_info = []  # [(label, start_col, end_col, color)]
    col_idx = 1
    for label, color, cols in _excel_groups():
        kept = [(h, g) for (h, g) in cols if h not in exclude]
        if not kept:
            continue
        group_info.append((label, col_idx, col_idx + len(kept) - 1, color))
        for header, getter in kept:
            all_cols.append((header, getter, color))
        col_idx += len(kept)
    if not all_cols:
        return jsonify({'error': "Nessun campo selezionato per l'export"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Alberi'

    hdr_font   = Font(bold=True)
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Row 1: group labels (merged)
    for label, start, end, color in group_info:
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=label)
        cell.fill      = PatternFill(fill_type='solid', fgColor=color)
        cell.font      = Font(bold=True, size=11)
        cell.alignment = center

    # Row 2: column headers
    for col_num, (header, _, color) in enumerate(all_cols, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill      = PatternFill(fill_type='solid', fgColor=color)
        cell.font      = hdr_font
        cell.alignment = center

    # Data rows
    for row_num, t in enumerate(trees, 3):
        for col_num, (_, getter, _color) in enumerate(all_cols, 1):
            val = getter(t)
            ws.cell(row=row_num, column=col_num, value=val if val != '' else None)

    # Column widths (approx by header length, capped)
    for col_num, (header, _, _) in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = max(10, min(35, len(header) + 3))

    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = 'C3'  # freeze first 2 cols + 2 header rows

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, download_name='alberi.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# Canonical column list shared by export and import.
# Format: (gpkg_column_name, sql_type, export_getter, tree_attr_or_None)
# tree_attr is the Tree model attribute name for import; None = skip on import.
# Multi-value fields (conflitti etc.) use a special prefix 'CSV:' handled below.
_GPKG_COLS = [
    ('id',                    'INTEGER', lambda t: t.id,                          None),
    ('custom_id',             'TEXT',    lambda t: t.custom_id,                   'custom_id'),
    ('city',                  'TEXT',    lambda t: t.city,                        None),  # city comes from the form
    ('address',               'TEXT',    lambda t: t.address or '',               'address'),
    ('latitude',              'REAL',    lambda t: t.latitude,                    'latitude'),
    ('longitude',             'REAL',    lambda t: t.longitude,                   'longitude'),
    ('species',               'TEXT',    lambda t: t.species,                     'species'),
    ('condition',             'TEXT',    lambda t: t.condition or '',             'condition'),
    ('age',                   'TEXT',    lambda t: t.age or '',                   'age'),
    ('cpc',                   'TEXT',    lambda t: t.cpc or '',                   'cpc'),
    ('location',              'TEXT',    lambda t: t.location or '',              'location'),
    ('height',                'TEXT',    lambda t: t.height or '',               'height'),
    ('dimora',                'TEXT',    lambda t: t.dimora or '',               'dimora'),
    ('stadio_sviluppo',       'TEXT',    lambda t: t.stadio_sviluppo or '',      'stadio_sviluppo'),
    ('posizione_sociale',     'TEXT',    lambda t: t.posizione_sociale or '',    'posizione_sociale'),
    ('localizzazione',        'TEXT',    lambda t: t.localizzazione or '',       'localizzazione'),
    ('vincoli',               'TEXT',    lambda t: t.vincoli or '',              'vincoli'),
    ('tree_height_m',         'REAL',    lambda t: t.tree_height_m,              'tree_height_m'),
    ('circonferenza_cm',      'REAL',    lambda t: t.circonferenza_cm,           'circonferenza_cm'),
    ('trunk_diameter_cm',     'REAL',    lambda t: t.trunk_diameter_cm,          'trunk_diameter_cm'),
    ('crown_diameter_m',      'REAL',    lambda t: t.crown_diameter_m,           'crown_diameter_m'),
    ('branch_diam_cm',        'REAL',    lambda t: t.branch_diam_cm,             'branch_diam_cm'),
    ('branch_length_m',       'REAL',    lambda t: t.branch_length_m,            'branch_length_m'),
    ('branch_height_m',       'REAL',    lambda t: t.branch_height_m,            'branch_height_m'),
    ('target_height_m',       'REAL',    lambda t: t.target_height_m,            'target_height_m'),
    ('post_tree_height_m',    'REAL',    lambda t: t.post_tree_height_m,         'post_tree_height_m'),
    ('post_circonferenza_cm', 'REAL',    lambda t: t.post_circonferenza_cm,      'post_circonferenza_cm'),
    ('post_branch_diam_cm',   'REAL',    lambda t: t.post_branch_diam_cm,        'post_branch_diam_cm'),
    ('post_branch_length_m',  'REAL',    lambda t: t.post_branch_length_m,       'post_branch_length_m'),
    ('post_branch_height_m',  'REAL',    lambda t: t.post_branch_height_m,       'post_branch_height_m'),
    ('post_target_height_m',  'REAL',    lambda t: t.post_target_height_m,       'post_target_height_m'),
    ('pericolo_rami',         'TEXT',    lambda t: t.pericolo_rami or '',        'pericolo_rami'),
    ('pericolo_tronco',       'TEXT',    lambda t: t.pericolo_tronco or '',      'pericolo_tronco'),
    ('pericolo_colletto',     'TEXT',    lambda t: t.pericolo_colletto or '',    'pericolo_colletto'),
    ('pericolo_zolla',        'TEXT',    lambda t: t.pericolo_zolla or '',       'pericolo_zolla'),
    ('bersaglio_chioma_tipo', 'TEXT',    lambda t: t.bersaglio_chioma_tipo or '', 'bersaglio_chioma_tipo'),
    ('bersaglio_chioma_value','TEXT',    lambda t: t.bersaglio_chioma_value or '','bersaglio_chioma_value'),
    ('bersaglio_chioma_flow', 'REAL',    lambda t: t.bersaglio_chioma_flow,      'bersaglio_chioma_flow'),
    ('bersaglio_chioma',      'INTEGER', lambda t: t.bersaglio_chioma,           'bersaglio_chioma'),
    ('bersaglio_ramo_tipo',   'TEXT',    lambda t: t.bersaglio_ramo_tipo or '',  'bersaglio_ramo_tipo'),
    ('bersaglio_ramo_value',  'TEXT',    lambda t: t.bersaglio_ramo_value or '', 'bersaglio_ramo_value'),
    ('bersaglio_ramo_flow',   'REAL',    lambda t: t.bersaglio_ramo_flow,        'bersaglio_ramo_flow'),
    ('bersaglio_ramo',        'INTEGER', lambda t: t.bersaglio_ramo,             'bersaglio_ramo'),
    ('moltiplicatore',        'INTEGER', lambda t: t.moltiplicatore,             'moltiplicatore'),
    ('monitoraggio',          'TEXT',    lambda t: t.monitoraggio or '',         'monitoraggio'),
    ('urgenza',               'TEXT',    lambda t: t.urgenza or '',              'urgenza'),
    ('next_check',            'TEXT',    lambda t: t.next_check.strftime('%Y-%m-%d') if t.next_check else '', 'next_check'),
    ('comments',              'TEXT',    lambda t: t.comments or '',             'comments'),
    ('actions',               'TEXT',    lambda t: t.actions or '',              'actions'),
    ('conflitti',             'TEXT',    None,                                    'CSV:conflitti_list'),
    ('agenti_carie',          'TEXT',    None,                                    'CSV:agenti_carie'),
    ('altri_patogeni',        'TEXT',    None,                                    'CSV:altri_patogeni'),
    ('prescrizioni_val',      'TEXT',    None,                                    'CSV:prescrizioni_val'),
    ('prescrizioni_mit',      'TEXT',    None,                                    'CSV:prescrizioni_mit'),
    ('prescrizioni_col',      'TEXT',    None,                                    'CSV:prescrizioni_col'),
]

def _gpkg_jlist(val):
    """DB JSON array → comma-separated string for GPKG storage."""
    if not val: return ''
    try:
        lst = json.loads(val) if isinstance(val, str) else val
        return ', '.join(str(x) for x in lst) if isinstance(lst, list) else str(lst)
    except: return ''

def _gpkg_csv_to_json(val):
    """Comma-separated GPKG string → DB JSON array string."""
    if not val or not str(val).strip(): return None
    items = [x.strip() for x in str(val).split(',') if x.strip()]
    return json.dumps(items, ensure_ascii=False) if items else None

@app.route('/export/gpkg/columns', methods=['GET'])
@auth_required
def export_gpkg_columns():
    """Nomi (chiavi) di default delle colonne del GPKG, per la finestra di rinomina in export."""
    return jsonify([{'name': n, 'type': typ} for n, typ, _, _ in _GPKG_COLS])


def _san_gpkg_ident(name):
    """Ripulisce un nome colonna in un identificatore SQL valido (o None se vuoto)."""
    s = re.sub(r'[^0-9A-Za-z_]', '_', str(name).strip())
    if not s:
        return None
    if s[0].isdigit():
        s = '_' + s
    return s


def _resolve_gpkg_names(overrides):
    """Restituisce la lista dei nomi colonna di output applicando le rinomine richieste.
    Garantisce nomi validi, non duplicati e diversi dalle colonne riservate (fid, geom)."""
    reserved = {'fid', 'geom'}
    out_names = []
    seen = {r.lower() for r in reserved}
    for (n, _typ, _g, _attr) in _GPKG_COLS:
        target = n
        if n in overrides:
            cand = _san_gpkg_ident(overrides[n])
            if cand and cand.lower() not in seen:
                target = cand
        base, k = target, 2
        while target.lower() in seen:
            target = f'{base}_{k}'
            k += 1
        seen.add(target.lower())
        out_names.append(target)
    return out_names


@app.route('/export/gpkg', methods=['GET'])
@auth_required
def export_gpkg():
    trees = [t for t in _scoped_tree_query().all()
             if t.latitude is not None and t.longitude is not None]

    def make_point(lon, lat):
        return b'GP\x00\x01' + struct.pack('<i', 4326) + struct.pack('<BIdd', 1, 1, lon, lat)

    # CSV multi-value getters (not in _GPKG_COLS lambdas to avoid closure issues)
    _csv_getters = {
        'conflitti':       lambda t: _gpkg_jlist(t.conflitti_list),
        'agenti_carie':    lambda t: _gpkg_jlist(t.agenti_carie),
        'altri_patogeni':  lambda t: _gpkg_jlist(t.altri_patogeni),
        'prescrizioni_val':lambda t: _gpkg_jlist(t.prescrizioni_val),
        'prescrizioni_mit':lambda t: _gpkg_jlist(t.prescrizioni_mit),
        'prescrizioni_col':lambda t: _gpkg_jlist(t.prescrizioni_col),
    }

    def get_val(col_name, getter, t):
        if getter is not None:
            return getter(t)
        return _csv_getters[col_name](t)

    # Rinomina delle chiavi (colonne) richiesta dall'utente: {nome_default: nome_nuovo}
    overrides = {}
    rename_param = request.args.get('rename')
    if rename_param:
        try:
            raw = json.loads(rename_param)
            if isinstance(raw, dict):
                overrides = {str(k): v for k, v in raw.items()}
        except Exception:
            overrides = {}
    out_names = _resolve_gpkg_names(overrides)

    # (nome_default, nome_output, tipo, getter) — il valore si estrae col nome default
    cols = [(n, out, typ, g) for (n, typ, g, _), out in zip(_GPKG_COLS, out_names)]

    # Esclusione dei campi non desiderati (per nome colonna di default)
    exclude = _export_exclude_set()
    if exclude:
        cols = [c for c in cols if c[0] not in exclude]
    if not cols:
        return jsonify({'error': "Nessun campo selezionato per l'export"}), 400

    col_defs = ', '.join(f'{out} {typ}' for _, out, typ, _ in cols)
    col_names = ', '.join(out for _, out, _, _ in cols)
    placeholders = ', '.join(['?'] * len(cols))

    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix='.gpkg', delete=False) as tmp:
            tmp_path = tmp.name

        conn = sqlite3.connect(tmp_path)
        cur  = conn.cursor()

        cur.execute('PRAGMA application_id = 1196444487')
        cur.execute('PRAGMA user_version   = 10200')
        conn.commit()

        cur.executescript(f"""
            CREATE TABLE gpkg_spatial_ref_sys (
                srs_name TEXT NOT NULL, srs_id INTEGER NOT NULL PRIMARY KEY,
                organization TEXT NOT NULL, organization_coordsys_id INTEGER NOT NULL,
                definition TEXT NOT NULL, description TEXT
            );
            INSERT INTO gpkg_spatial_ref_sys VALUES
                ('Undefined cartesian SRS', -1, 'NONE', -1, 'undefined', NULL),
                ('Undefined geographic SRS', 0, 'NONE', 0, 'undefined', NULL),
                ('WGS 84 geodetic', 4326, 'EPSG', 4326,
                 'GEOGCS["WGS 84",DATUM["WGS_1984",SPHEROID["WGS 84",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["degree",0.0174532925199433]]',
                 'WGS 84');
            CREATE TABLE gpkg_contents (
                table_name TEXT NOT NULL PRIMARY KEY, data_type TEXT NOT NULL,
                identifier TEXT, description TEXT DEFAULT '',
                last_change DATETIME NOT NULL DEFAULT (strftime('%Y-%m-%dT%H:%M:%SZ','now')),
                min_x REAL, min_y REAL, max_x REAL, max_y REAL, srs_id INTEGER
            );
            CREATE TABLE gpkg_geometry_columns (
                table_name TEXT NOT NULL, column_name TEXT NOT NULL,
                geometry_type_name TEXT NOT NULL, srs_id INTEGER NOT NULL,
                z TINYINT NOT NULL, m TINYINT NOT NULL,
                CONSTRAINT pk_geom_cols PRIMARY KEY (table_name, column_name)
            );
            CREATE TABLE alberi (
                fid INTEGER PRIMARY KEY AUTOINCREMENT,
                geom BLOB,
                {col_defs}
            );
        """)

        lons = [t.longitude for t in trees]
        lats = [t.latitude  for t in trees]
        cur.execute(
            "INSERT INTO gpkg_contents (table_name,data_type,identifier,srs_id,min_x,min_y,max_x,max_y) "
            "VALUES ('alberi','features','alberi',4326,?,?,?,?)",
            (min(lons) if lons else None, min(lats) if lats else None,
             max(lons) if lons else None, max(lats) if lats else None)
        )
        cur.execute("INSERT INTO gpkg_geometry_columns VALUES ('alberi','geom','POINT',4326,0,0)")

        INSERT = f"INSERT INTO alberi (geom, {col_names}) VALUES (?, {placeholders})"
        for t in trees:
            cur.execute(INSERT,
                [make_point(t.longitude, t.latitude)] +
                [get_val(n, g, t) for n, _out, _typ, g in cols])

        conn.commit()
        conn.close()

        buf = io.BytesIO()
        with open(tmp_path, 'rb') as f:
            buf.write(f.read())
        buf.seek(0)
        return send_file(buf, download_name='alberi.gpkg', as_attachment=True,
                         mimetype='application/geopackage+sqlite3')
    finally:
        if tmp_path:
            try: os.unlink(tmp_path)
            except Exception: pass

# -----------------------
# Report — Schede albero (PLACEHOLDER)
# -----------------------

def _scoped_tree_query():
    """Query alberi filtrata per ruolo del chiamante e per il parametro `ids`.
    Stessa logica di visibilità/selezione usata dagli export (excel/gpkg)."""
    role, user_id, user_city = (request.user.get(k) for k in ('role', 'user_id', 'city'))
    query = Tree.query
    if role == 'user':   query = query.filter(Tree.owner_id == user_id)
    elif role == 'city': query = city_tree_filter(query, user_id, user_city)
    elif role == 'superuser':
        did = _demo_user_id()
        if did and did != user_id: query = query.filter(Tree.owner_id != did)
    ids_param = request.args.get('ids')
    if ids_param:
        id_list = [int(x) for x in ids_param.split(',') if x.strip().isdigit()]
        query = query.filter(Tree.id.in_(id_list))
    return query


@app.route('/report/templates', methods=['GET'])
@auth_required
def report_templates_list():
    """Elenco dei template di scheda albero disponibili (per il frontend)."""
    return jsonify(report_tpl.list_templates())


@app.route('/report/scheda', methods=['GET'])
@auth_required
def report_scheda():
    """Genera il report "schede albero" per gli alberi visibili/selezionati.

    Parametri:
      - ids:      selezione opzionale (`1,2,3`); assente = tutti i visibili
      - template: id del template (assente = template di default)

    Il template di default ("scheda_arete") produce l'.xlsx identico al template
    ufficiale ARETE (foglio ORD), compilato coi dati del database — un file per
    albero, oppure uno .zip se gli alberi selezionati sono più d'uno.
    """
    template_id = request.args.get('template')
    template = report_tpl.get_template(template_id)
    if template is None:
        return jsonify({'error': f'Template non trovato: {template_id}'}), 404

    trees = _scoped_tree_query().all()
    if not trees:
        return jsonify({'error': 'Nessun albero da includere nelle schede'}), 404

    filename, content, mimetype = report_tpl.render_schede(trees, template)

    buf = io.BytesIO(content)
    buf.seek(0)
    return send_file(buf, download_name=filename, as_attachment=True,
                     mimetype=mimetype)

# -----------------------
# Import GPKG
# -----------------------

def _gpkg_parse_float(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '.').strip())
    except (ValueError, TypeError):
        return None

def _gpkg_parse_date(val):
    if not val:
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%m/%y'):
        try:
            return datetime.strptime(str(val).strip(), fmt).date()
        except (ValueError, AttributeError):
            continue
    return None

def _gpkg_point(blob):
    """Decode a GeoPackage binary geometry blob and return (lon, lat)."""
    if not blob or len(blob) < 8:
        return None, None
    try:
        if bytes(blob[:2]) != b'GP':
            return None, None
        flags = blob[3]
        env_code = (flags >> 1) & 0x07
        env_size = {0: 0, 1: 32, 2: 48, 3: 48, 4: 64}.get(env_code, 0)
        wkb = bytes(blob[8 + env_size:])
        if len(wkb) < 21:
            return None, None
        endian = '<' if wkb[0] == 1 else '>'
        x, y = struct.unpack(f'{endian}dd', wkb[5:21])
        return x, y
    except Exception:
        return None, None

def _reverse_geocode_comune(lat, lon):
    """Return Italian comune name from coordinates via Nominatim (1 call per import)."""
    url = (f'https://nominatim.openstreetmap.org/reverse'
           f'?lat={lat}&lon={lon}&format=json&zoom=10&addressdetails=1&accept-language=it')
    req = urllib.request.Request(url, headers={'User-Agent': 'SilvaePro/1.0 (tree-inventory)'})
    try:
        with urllib.request.urlopen(req, timeout=6) as r:
            data = json.loads(r.read().decode())
        addr = data.get('address', {})
        return (addr.get('city') or addr.get('town') or
                addr.get('village') or addr.get('municipality') or
                addr.get('county') or '').strip()
    except Exception:
        return ''

# ---------- GPKG import: field-detection helpers ----------

_GPKG_FIELD_CANDIDATES = {
    'custom_id':          ['custom_id', 'name', '_numero', '_name', 'id_albero', 'numero'],
    'species':            ['species', '_tassonomi', 'specie', 'nome_scientifico', 'taxon'],
    'species_ita':        ['_nome ital', '_nome_ital', 'nome_italiano', 'nome_comune'],
    'condition':          ['condition', 'condizione', '_classe vt', '_classe_vt'],
    'cpc':                ['cpc', '_classe vt', '_classe_vt'],
    'age':                ['age', '_f. fisiol', '_f_fisiol'],
    'address':            ['address', "_localita'", "localita'", '_localita', 'localita', 'indirizzo', 'via'],
    'height':             ['height', '_altezza', 'altezza', 'h_albero'],
    'crown_diameter_m':   ['crown_diameter_m', '_diametro', 'diametro_chioma'],
    'circonferenza_cm':   ['circonferenza_cm', '_circonfer', '_circonferenza', 'circonferenza'],
    'localizzazione':     ['localizzazione'],
    'location':           ['location'],
    'stazione_raw':       ['_stazione', 'stazione'],
    'next_check':         ['next_check', '_da contro', '_da_contro', 'prossima_ispezione', 'data_controllo'],
    'next_check_fb':      ['_rivedere'],
    'longitude':          ['longitude', 'xcoord', 'x', 'lon'],
    'latitude':           ['latitude', 'ycoord', 'y', 'lat'],
    'geom':               ['geom', 'geometry', 'the_geom', 'shape'],
    'dimora':             ['dimora'],
    'stadio_sviluppo':    ['stadio_sviluppo'],
    'posizione_sociale':  ['posizione_sociale'],
    'vincoli':            ['vincoli'],
    'tree_height_m':      ['tree_height_m'],
    'trunk_diameter_cm':  ['trunk_diameter_cm'],
    'branch_diam_cm':     ['branch_diam_cm'],
    'branch_length_m':    ['branch_length_m'],
    'branch_height_m':    ['branch_height_m'],
    'target_height_m':    ['target_height_m'],
    'post_tree_height_m':    ['post_tree_height_m'],
    'post_circonferenza_cm': ['post_circonferenza_cm'],
    'post_branch_diam_cm':   ['post_branch_diam_cm'],
    'post_branch_length_m':  ['post_branch_length_m'],
    'post_branch_height_m':  ['post_branch_height_m'],
    'post_target_height_m':  ['post_target_height_m'],
    'pericolo_rami':      ['pericolo_rami'],
    'pericolo_tronco':    ['pericolo_tronco'],
    'pericolo_colletto':  ['pericolo_colletto'],
    'pericolo_zolla':     ['pericolo_zolla'],
    'bersaglio_chioma_tipo':  ['bersaglio_chioma_tipo'],
    'bersaglio_chioma_value': ['bersaglio_chioma_value'],
    'bersaglio_chioma_flow':  ['bersaglio_chioma_flow'],
    'bersaglio_chioma':       ['bersaglio_chioma'],
    'bersaglio_ramo_tipo':    ['bersaglio_ramo_tipo'],
    'bersaglio_ramo_value':   ['bersaglio_ramo_value'],
    'bersaglio_ramo_flow':    ['bersaglio_ramo_flow'],
    'bersaglio_ramo':         ['bersaglio_ramo'],
    'moltiplicatore':     ['moltiplicatore'],
    'monitoraggio':       ['monitoraggio'],
    'urgenza':            ['urgenza'],
    'comments':           ['comments'],
    'actions':            ['actions'],
    'conflitti_list':     ['conflitti'],
    'agenti_carie':       ['agenti_carie'],
    'altri_patogeni':     ['altri_patogeni'],
    'prescrizioni_val':   ['prescrizioni_val'],
    'prescrizioni_mit':   ['prescrizioni_mit'],
    'prescrizioni_col':   ['prescrizioni_col'],
}

def _gpkg_detect_C(cols):
    """Auto-detect field mapping from GPKG column names."""
    lc = {c.lower(): c for c in cols}
    C = {}
    for field, candidates in _GPKG_FIELD_CANDIDATES.items():
        for cand in candidates:
            if cand.lower() in lc:
                C[field] = lc[cand.lower()]
                break
        else:
            C[field] = None
    return C

def _gpkg_user_C(cols, user_mapping):
    """Build field mapping from explicit user config {gpkg_col: silvae_field}."""
    col_set = set(cols)
    C = {f: None for f in _GPKG_FIELD_CANDIDATES}
    for gpkg_col, silvae_field in user_mapping.items():
        if silvae_field and silvae_field in C and gpkg_col in col_set:
            C[silvae_field] = gpkg_col
    return C

def _gpkg_inspect_mapping(cols):
    """Return user-facing auto-mapping for the inspect endpoint: {gpkg_col: silvae_field}."""
    C = _gpkg_detect_C(cols)
    internal = {'stazione_raw', 'next_check_fb', 'geom'}
    result = {}
    for field, gpkg_col in C.items():
        if field in internal or not gpkg_col:
            continue
        if gpkg_col not in result:
            result[gpkg_col] = field
    if C.get('stazione_raw') and C['stazione_raw'] not in result:
        result[C['stazione_raw']] = 'location'
    if C.get('next_check_fb') and C['next_check_fb'] not in result:
        result[C['next_check_fb']] = 'next_check'
    return result

def _set_geom(tree):
    """Write PostGIS geom from lat/lon — silently skipped if PostGIS is unavailable."""
    if not tree.id or not tree.latitude or not tree.longitude:
        return
    try:
        db.session.execute(
            text("UPDATE tree SET geom = ST_SetSRID(ST_MakePoint(:lon,:lat),4326) WHERE id=:id"),
            {'lon': tree.longitude, 'lat': tree.latitude, 'id': tree.id}
        )
    except Exception:
        db.session.rollback()

def _gpkg_read_table(f_or_path):
    """Open a GPKG file object or path, return (cols, rows, table_name) or raise."""
    import_path = None
    try:
        if hasattr(f_or_path, 'read'):
            with tempfile.NamedTemporaryFile(suffix='.gpkg', delete=False) as tmp:
                f_or_path.save(tmp.name)
                import_path = tmp.name
            path = import_path
        else:
            path = f_or_path
        src = sqlite3.connect(path)
        src.row_factory = sqlite3.Row
        cur = src.cursor()
        try:
            cur.execute("SELECT table_name FROM gpkg_contents WHERE data_type='features' LIMIT 1")
            r = cur.fetchone()
            table_name = r[0] if r else None
        except Exception:
            table_name = None
        if not table_name:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table' "
                        "AND name NOT LIKE 'gpkg_%' AND name NOT LIKE 'rtree_%' "
                        "AND name NOT LIKE 'sqlite_%'")
            rows_t = cur.fetchall()
            table_name = rows_t[0][0] if rows_t else None
        if not table_name:
            src.close()
            raise ValueError('Nessuna tabella features trovata nel file GPKG')
        cur.execute(f"PRAGMA table_info('{table_name}')")
        cols = [r[1] for r in cur.fetchall()]
        cur.execute(f"SELECT * FROM '{table_name}'")
        rows = [dict(r) for r in cur.fetchall()]
        src.close()
        return cols, rows, table_name
    finally:
        if import_path:
            try: os.unlink(import_path)
            except Exception: pass

@app.route('/import/gpkg/inspect', methods=['POST'])
@auth_required
def inspect_gpkg_route():
    f = request.files.get('file')
    if not f:
        return jsonify({'message': 'Nessun file inviato'}), 400
    try:
        cols, rows, _ = _gpkg_read_table(f)
    except Exception as e:
        return jsonify({'message': f'Errore lettura GPKG: {e}'}), 400

    skip = {'fid', 'geom', 'geometry', 'the_geom', 'shape'}
    display_cols = [c for c in cols if c.lower() not in skip]
    sample = {}
    if rows:
        for k in display_cols:
            v = rows[0].get(k)
            if not isinstance(v, (bytes, bytearray)):
                sample[k] = str(v) if v is not None else ''
    return jsonify({
        'columns':      display_cols,
        'auto_mapping': _gpkg_inspect_mapping(cols),
        'sample':       sample,
    })

@app.route('/import/gpkg', methods=['POST'])
@auth_required
def import_gpkg_route():
    user_id   = request.user.get('user_id')
    role      = request.user.get('role')
    user_city = request.user.get('city')

    if 'file' not in request.files:
        return jsonify({'message': 'Nessun file inviato'}), 400

    f           = request.files['file']
    city        = (request.form.get('city') or '').strip()
    on_conflict = request.form.get('on_conflict', 'skip')

    if not city:
        if role in ('user', 'city') and user_city:
            city = user_city
        # else: city will be auto-detected from coordinates after parsing the file

    if not f.filename:
        return jsonify({'message': 'Nessun file selezionato'}), 400

    try:
        cols, rows, _ = _gpkg_read_table(f)
    except Exception as e:
        return jsonify({'message': f'Errore lettura GPKG: {e}'}), 400

    # --- Column detection ------------------------------------------------
    mapping_json = request.form.get('mapping')
    if mapping_json:
        try:
            C = _gpkg_user_C(cols, json.loads(mapping_json))
        except (ValueError, TypeError):
            return jsonify({'message': 'Mapping JSON non valido'}), 400
    else:
        C = _gpkg_detect_C(cols)

    # --- Row helpers -------------------------------------------------------
    def gs(key):
        col = C.get(key)
        if not col: return None
        v = row.get(col)
        return str(v).strip() or None if v is not None else None

    def gf(key):
        col = C.get(key)
        return _gpkg_parse_float(row.get(col)) if col else None

    def gi(key):
        col = C.get(key)
        if not col: return None
        try: return int(str(row.get(col)).strip())
        except: return None

    def gcsvjson(key):
        col = C.get(key)
        return _gpkg_csv_to_json(row.get(col)) if col else None

    # --- Apply all detected fields to a Tree instance ----------------------
    def apply_fields(tree):
        if gf('latitude') is not None:
            tree.latitude  = gf('latitude')
            tree.longitude = gf('longitude')

        sp = gs('species') or gs('species_ita') or 'Sconosciuta'
        tree.species = sp

        tree.condition = gs('condition') or '—'

        for key in ('address','height','age','cpc','location',
                    'dimora','stadio_sviluppo','posizione_sociale',
                    'localizzazione','vincoli',
                    'pericolo_rami','pericolo_tronco',
                    'pericolo_colletto','pericolo_zolla',
                    'bersaglio_chioma_tipo','bersaglio_chioma_value',
                    'bersaglio_ramo_tipo','bersaglio_ramo_value',
                    'monitoraggio','urgenza','comments','actions'):
            v = gs(key)
            if v is not None:
                setattr(tree, key, v)
        if tree.cpc and tree.cpc.upper() == 'ABBATTUO':
            tree.cpc = 'ABBATTUTO'
        if tree.cpc and (not tree.condition or tree.condition == '—'):
            tree.condition = tree.cpc

        staz = gs('stazione_raw')
        if staz and not gs('localizzazione') and not gs('location'):
            localiz_vals = [v.lower() for v in dd.localizzazione]
            if staz.lower() in localiz_vals:
                tree.localizzazione = staz
            else:
                tree.location = staz

        for key in ('tree_height_m','circonferenza_cm','trunk_diameter_cm',
                    'crown_diameter_m','branch_diam_cm','branch_length_m',
                    'branch_height_m','target_height_m',
                    'post_tree_height_m','post_circonferenza_cm',
                    'post_branch_diam_cm','post_branch_length_m',
                    'post_branch_height_m','post_target_height_m',
                    'bersaglio_chioma_flow','bersaglio_ramo_flow'):
            v = gf(key)
            if v is not None:
                setattr(tree, key, v)

        for key in ('bersaglio_chioma','bersaglio_ramo','moltiplicatore'):
            v = gi(key)
            if v is not None:
                setattr(tree, key, v)

        for key in ('conflitti_list','agenti_carie','altri_patogeni',
                    'prescrizioni_val','prescrizioni_mit','prescrizioni_col'):
            v = gcsvjson(key)
            if v is not None:
                setattr(tree, key, v)

        nd = _gpkg_parse_date(gs('next_check')) or _gpkg_parse_date(gs('next_check_fb'))
        if nd is not None:
            tree.next_check = nd

    # --- Auto-detect comune from first valid coordinate if city not provided ---
    city_autodetected = False
    if not city:
        for row in rows:
            lat = lon = None
            if C['latitude'] and C['longitude']:
                lat = _gpkg_parse_float(row.get(C['latitude']))
                lon = _gpkg_parse_float(row.get(C['longitude']))
            if (lat is None or lon is None) and C['geom']:
                blob = row.get(C['geom'])
                if blob:
                    lon, lat = _gpkg_point(blob)
            if lat is not None and lon is not None:
                city = _reverse_geocode_comune(lat, lon)
                city_autodetected = bool(city)
                break
        if not city:
            return jsonify({'message': 'Comune non rilevabile dalle coordinate. Inseriscilo manualmente.'}), 400

    # --- Main loop ---------------------------------------------------------
    inserted = skipped = errors = 0
    error_details = []

    for row in rows:
        # Coordinates: prefer explicit x/y columns, fall back to geometry blob
        lon = lat = None
        if C['longitude'] and C['latitude']:
            lon = _gpkg_parse_float(row.get(C['longitude']))
            lat = _gpkg_parse_float(row.get(C['latitude']))
        if (lon is None or lat is None) and C['geom']:
            blob = row.get(C['geom'])
            if blob:
                lon, lat = _gpkg_point(blob)

        # Override row coords so apply_fields() picks them up
        if lon is not None and C['longitude']:
            row[C['longitude']] = lon
        if lat is not None and C['latitude']:
            row[C['latitude']] = lat

        custom_id = str(row.get(C['custom_id']) or '').strip() if C['custom_id'] else ''
        if not custom_id:
            skipped += 1
            continue

        try:
            # Deduplicate by coordinates (trees don't move); fall back to
            # custom_id+city only when coordinates are unavailable.
            existing = None
            if lat is not None and lon is not None:
                tol = 1e-6
                existing = Tree.query.filter(
                    Tree.city == city,
                    Tree.latitude.between(lat - tol, lat + tol),
                    Tree.longitude.between(lon - tol, lon + tol)
                ).first()
            else:
                existing = Tree.query.filter_by(custom_id=custom_id, city=city).first()

            if existing:
                if on_conflict == 'update':
                    existing.custom_id = custom_id
                    existing.owner_id  = user_id
                    apply_fields(existing)
                    db.session.commit()
                    inserted += 1
                else:
                    skipped += 1
                continue

            tree = Tree(custom_id=custom_id, city=city, owner_id=user_id,
                        species='Sconosciuta', condition='—')
            apply_fields(tree)
            db.session.add(tree)
            db.session.commit()
            inserted += 1
        except Exception as e:
            db.session.rollback()
            errors += 1
            error_details.append(f"fid={row.get('fid','?')} id={custom_id}: {e}")

    return jsonify({'inserted': inserted, 'skipped': skipped, 'errors': errors,
                    'total': len(rows), 'city': city, 'city_autodetected': city_autodetected,
                    'error_details': error_details})

# -----------------------
# Voice intent (Groq NLU fallback)
# -----------------------
GROQ_SYSTEM = """Sei un assistente per un'app di censimento alberi urbani.
Estrai intent e parametri dalla frase italiana dell'utente.
Intent possibili:
  OPEN_TREE       params: { customId }
  EDIT_TREE       params: { customId }
  OPEN_HISTORY    params: { customId }
  LOCATE_TREE     params: { customId }
  NEARBY_RADIUS   params: { radius }   (radius in metres, integer)
  NEARBY_TOGGLE   params: {}
  FILTER_ADDRESS   params: { address }  (street name or locality to search, e.g. "Pollenzo" or "via Roma")
  FILTER_CONDITION params: { label }   (Ottimo|Buono|Discreto|Scarso|Critico|Morto)
  RESET_FILTERS   params: {}
  ADD_TREE        params: {}
  SWITCH_MAP      params: {}
  SWITCH_LIST     params: {}
  COUNT_TREES     params: {}
Rispondi SOLO con JSON valido: {"intent":"...","params":{...}}
Se non riesci a determinare l'intent rispondi: {"intent":null,"params":{}}"""

@app.route('/api/voice_intent', methods=['POST'])
@auth_required
def voice_intent():
    data = request.get_json()
    transcript = (data or {}).get('transcript', '').strip()
    if not transcript:
        return jsonify({'intent': None, 'params': {}}), 200

    api_key = os.environ.get('GROQ_API_KEY')
    if not api_key:
        return jsonify({'error': 'GROQ_API_KEY not set'}), 503

    try:
        client = GroqClient(api_key=api_key)
        resp = client.chat.completions.create(
            model='llama-3.1-8b-instant',
            messages=[
                {'role': 'system', 'content': GROQ_SYSTEM},
                {'role': 'user',   'content': transcript},
            ],
            temperature=0,
            max_tokens=80,
        )
        raw = resp.choices[0].message.content.strip()
        result = json.loads(raw)
        return jsonify(result), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# -----------------------
# Frontend static serving
# -----------------------
@app.route('/')
def serve_index():
    return send_from_directory('frontend', 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    return send_from_directory('frontend', path)

# -----------------------
# Bootstrap
# -----------------------
with app.app_context():
    if not User.query.filter_by(role='superuser').first():
        su = User(username='admin', role='superuser', city=None)
        su.set_password('admin')
        db.session.add(su); db.session.commit()
        print("Created default superuser: admin / admin")
    # Ensure the demo accounts exist and start from a clean demo state.
    try:
        reset_demo()
        print(f"Demo accounts ready: {DEMO_USERNAME} / {DEMO_PASSWORD}"
              f"  |  {DEMO_CITY_USERNAME} / {DEMO_CITY_PASSWORD}")
    except Exception as e:
        print(f"[DEMO] Seeding skipped: {e}")

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')
