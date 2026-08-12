"""Generazione della "Scheda di rilevamento ARETE" (foglio ORD) dal database.

Idea: invece di reimpaginare la scheda, si **riparte dal template ufficiale**
``Schede_Rilevamento_ARETE_DEMO_ver.2.0.xlsm`` (foglio ``ORD``) e si scrivono i
valori presi dal DB nelle celle giuste. Il layout, le etichette, le celle unite
e gli stili restano identici all'originale.

Requisiti del committente:
  - tabella **identica** a quella del template;
  - **niente macro** (l'output è .xlsx, il VBA viene scartato);
  - i **contenuti** arrivano direttamente dal database (comprese le celle che
    nell'originale erano formule: dove il DB ha già il risultato calcolato —
    p.es. il rischio — lo scriviamo come valore, così non servono formule).

``CELL_MAP`` è la mappa campo→cella: è il punto in cui correggere/estendere se
qualche valore va in una cella diversa.
"""

from __future__ import annotations

import io
import json
import os
import shutil
import subprocess
import tempfile
import zipfile
from datetime import date, datetime

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font

# --- Percorso del template -----------------------------------------------------
_PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(
    _PROJECT_ROOT,
    "Schede_Rilevamento_ARETE",
    "Schede_Rilevamento_ARETE_DEMO_ver.2.0.xlsm",
)
SHEET_NAME = "ORD"

# Logo SilvaePro in alto a sinistra (al posto della scritta "demo ver.2.0", C4).
LOGO_PATH = os.path.join(_PROJECT_ROOT, "frontend", "images", "logo_reduced.png")
LOGO_ANCHOR = "B2"   # angolo in alto a sinistra (riga 2 = inizio area di stampa)
LOGO_SIZE_PX = 62
_LOGO_BYTES: bytes | None = None

# Logo Protocollo ARETE in alto a destra (al posto della scritta "VERSIONE DEMO").
ARETE_LOGO_PATH = os.path.join(_PROJECT_ROOT, "frontend", "images", "arete_logo.png")
ARETE_ANCHOR = "O2"
ARETE_W_PX, ARETE_H_PX = 101, 46   # mantiene le proporzioni 214x97
_ARETE_BYTES: bytes | None = None

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ZIP_MIME = "application/zip"

# LibreOffice per la conversione xlsx -> PDF (deve essere installato).
SOFFICE_BIN = shutil.which("libreoffice") or shutil.which("soffice")
# Cartella di lavoro temporanea: sotto la root del progetto così è accessibile
# anche quando LibreOffice è installato come snap (che non può leggere /tmp).
WORK_BASE = os.environ.get("SCHEDA_WORK_DIR") or os.path.join(_PROJECT_ROOT, ".scheda_work")

# Byte del template caricati una sola volta (poi si ricarica da memoria per albero)
_TEMPLATE_BYTES: bytes | None = None


def _template_bytes() -> bytes:
    global _TEMPLATE_BYTES
    if _TEMPLATE_BYTES is None:
        with open(TEMPLATE_PATH, "rb") as fh:
            _TEMPLATE_BYTES = fh.read()
    return _TEMPLATE_BYTES


def _logo_bytes() -> bytes | None:
    """Logo SilvaePro già ridimensionato (PNG piccolo), per non appesantire i file.
    L'originale è 1024x1024: lo riduciamo a ~2x la dimensione di visualizzazione."""
    global _LOGO_BYTES
    if _LOGO_BYTES is None:
        try:
            from PIL import Image as PILImage
            im = PILImage.open(LOGO_PATH).convert("RGBA")
            # ritaglia il bordo trasparente così il marchio riempie il riquadro
            bbox = im.getbbox()
            if bbox:
                im = im.crop(bbox)
            im.thumbnail((LOGO_SIZE_PX * 2, LOGO_SIZE_PX * 2))
            buf = io.BytesIO()
            im.save(buf, format="PNG")   # RGBA: sfondo trasparente preservato
            _LOGO_BYTES = buf.getvalue()
        except (OSError, ImportError):
            _LOGO_BYTES = b""   # assente/non processabile: si prosegue senza logo
    return _LOGO_BYTES or None


def _arete_bytes() -> bytes | None:
    global _ARETE_BYTES
    if _ARETE_BYTES is None:
        try:
            with open(ARETE_LOGO_PATH, "rb") as fh:
                _ARETE_BYTES = fh.read()
        except OSError:
            _ARETE_BYTES = b""
    return _ARETE_BYTES or None


def _setup_id_cell(ws, custom_id) -> None:
    """L'ID sta nell'ultima colonna (U), troppo stretta e al bordo dell'area di
    stampa: LibreOffice non ne rispetta la larghezza. Gli do spazio unendo la
    casella su S6:U6 (l'area 'CODICE' non è più usata) con l'etichetta 'ID' in R6."""
    ws["R6"] = "ID"        # etichetta (al posto di 'CODICE', non più calcolato)
    ws["T6"] = None        # via la vecchia etichetta 'ID'
    ws["U6"] = None
    ws.merge_cells("S6:U6")
    ws["S6"] = "" if custom_id is None else str(custom_id)
    f = ws["S6"].font
    ws["S6"].font = Font(name=f.name, size=11, bold=True, color=f.color)
    ws["S6"].alignment = Alignment(horizontal="center", vertical="center")


def _add_logos(ws) -> None:
    """Logo SilvaePro in alto a sinistra (al posto di 'demo ver.2.0', C4) e logo
    Protocollo ARETE in alto a destra (al posto di 'VERSIONE DEMO', P3)."""
    ws["C4"] = None
    ws["P3"] = None   # via la scritta "VERSIONE DEMO"

    data = _logo_bytes()
    if data:
        img = XLImage(io.BytesIO(data))
        scale = LOGO_SIZE_PX / max(img.width, img.height)  # riquadro, proporzioni intatte
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, LOGO_ANCHOR)

    arete = _arete_bytes()
    if arete:
        aimg = XLImage(io.BytesIO(arete))
        aimg.width, aimg.height = ARETE_W_PX, ARETE_H_PX
        ws.add_image(aimg, ARETE_ANCHOR)


# --- Helpers per estrarre i valori dal Tree -----------------------------------

def _as_list(raw) -> list[str]:
    """Campi multi-valore: nel DB sono JSON (``["a","b"]``) o testo libero."""
    if raw is None:
        return []
    if isinstance(raw, (list, tuple)):
        return [str(x) for x in raw if str(x).strip()]
    s = str(raw).strip()
    if not s:
        return []
    try:
        val = json.loads(s)
        if isinstance(val, list):
            return [str(x) for x in val if str(x).strip()]
    except (ValueError, TypeError):
        pass
    # fallback: separatori comuni
    for sep in ("\n", ";", ","):
        if sep in s:
            return [p.strip() for p in s.split(sep) if p.strip()]
    return [s]


def _fmt_date(v) -> str:
    if isinstance(v, (date, datetime)):
        return v.strftime("%d/%m/%Y")
    return "" if v is None else str(v)


def _rischio(tree) -> dict:
    raw = getattr(tree, "rischio", None)
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        return json.loads(raw) or {}
    except (ValueError, TypeError):
        return {}


def _comp(section: dict, key: str) -> dict:
    """Sotto-dizionario di un componente (rami/tronco/colletto/zolla)."""
    v = section.get(key) if isinstance(section, dict) else None
    return v if isinstance(v, dict) else {}


def _round1(v):
    try:
        return round(float(v), 1)
    except (TypeError, ValueError):
        return v


def _risk_cell(comp: dict) -> str:
    """Testo per la colonna RISCHIO (R): rapporto + descrizione."""
    if not comp:
        return ""
    ratio = comp.get("risk_ratio") or comp.get("risk_ratio_1bers") or ""
    desc = comp.get("risk_description") or ""
    return " — ".join(p for p in (str(ratio), str(desc)) if p and p != "None")


# --- Scrittura di un blocco DIAGNOSI (una voce per riga) ----------------------

def _write_column_block(ws, col: str, first_row: int, last_row: int, values: list[str]):
    """Scrive ``values`` in verticale nella colonna ``col`` da ``first_row`` a
    ``last_row`` (le celle sono unite orizzontalmente: si scrive nell'ancora)."""
    for i, row in enumerate(range(first_row, last_row + 1)):
        ws[f"{col}{row}"] = values[i] if i < len(values) else None


def _fill_ord_sheet(ws, tree) -> None:
    """Riempie il foglio ORD con i dati di un singolo albero."""
    risk = _rischio(tree)
    attuale = risk.get("attuale", {}) if isinstance(risk, dict) else {}
    residuo = risk.get("residuo", {}) if isinstance(risk, dict) else {}

    # --- DATI GENERALI ---
    _set(ws, "J6", getattr(tree, "species", None))          # SPECIE
    # ID: gestito a parte con celle unite (vedi _setup_id_cell) per avere spazio.
    _set(ws, "H6", "")                                      # DATA (non nel Tree)
    _set(ws, "C7", getattr(tree, "dimora", None))           # DIMORA
    _set(ws, "G7", getattr(tree, "localizzazione", None))   # LOCALIZ
    _set(ws, "J7", getattr(tree, "posizione_sociale", None))# P.Soc.
    _set(ws, "N7", getattr(tree, "stadio_sviluppo", None))  # STADIO
    _set(ws, "S7", getattr(tree, "vincoli", None))          # Vincoli
    _set(ws, "D8", getattr(tree, "address", None)
         or getattr(tree, "location", None))                # UBICAZ
    conflitti = _as_list(getattr(tree, "conflitti_list", None))
    _set(ws, "L8", conflitti[0] if len(conflitti) > 0 else None)  # CONFLITTI 1
    _set(ws, "N8", conflitti[1] if len(conflitti) > 1 else None)  # CONFLITTI 2

    # --- DIMENSIONI (riga 9) ---
    _set(ws, "C9", getattr(tree, "tree_height_m", None))    # H
    diam = getattr(tree, "trunk_diameter_cm", None)
    if diam in (None, "") and getattr(tree, "circonferenza_cm", None):
        try:
            diam = round(float(tree.circonferenza_cm) / 3.141592653589793, 1)
        except (TypeError, ValueError):
            diam = None
    _set(ws, "E9", diam)                                    # D tr
    _set(ws, "H9", getattr(tree, "circonferenza_cm", None)) # Circ (era formula)
    _set(ws, "J9", getattr(tree, "crown_diameter_m", None)) # D ch
    _set(ws, "M9", getattr(tree, "branch_diam_cm", None))   # D br
    _set(ws, "O9", getattr(tree, "branch_length_m", None))  # L br
    _set(ws, "R9", getattr(tree, "branch_height_m", None))  # H br
    _set(ws, "T9", getattr(tree, "target_height_m", None))  # H bers

    # --- VALORE ECOLOGICO (riga 10, erano formule) ---
    _set(ws, "D10", getattr(tree, "valore_ecologico", None))  # Valore ecologico (€)
    _set(ws, "G10", getattr(tree, "bio_kg", None))            # Bio (kg)
    _set(ws, "J10", getattr(tree, "co2_kg_anno", None))       # CO2 (kg/anno)
    _set(ws, "M10", getattr(tree, "o2_kg_anno", None))        # O2 (kg/y)
    _set(ws, "O10", getattr(tree, "ia_kg_anno", None))        # I.A. (kg/y)

    # --- CONDIZIONI DI SALUTE / fitopatie ---
    _set(ws, "E12", getattr(tree, "condizione_salute_ecologica", None))
    _set(ws, "R12", ", ".join(_as_list(getattr(tree, "altri_patogeni", None))))
    _set(ws, "R13", ", ".join(_as_list(getattr(tree, "agenti_carie", None))))

    # --- DIAGNOSI (griglia righe 15-24) ---
    _write_column_block(ws, "B", 15, 24, _as_list(getattr(tree, "diag_zolla", None)))
    _write_column_block(ws, "G", 15, 24, _as_list(getattr(tree, "diag_colletto", None)))
    _write_column_block(ws, "L", 15, 19, _as_list(getattr(tree, "diag_fusto", None)))
    _write_column_block(ws, "L", 21, 24, _as_list(getattr(tree, "diag_castello", None)))
    _write_column_block(ws, "Q", 15, 19, _as_list(getattr(tree, "diag_ramificazione", None)))
    _write_column_block(ws, "Q", 21, 24, _as_list(getattr(tree, "diag_chioma", None)))

    # --- GRADO DI PERICOLO (classi dal DB) ---
    _set(ws, "E26", getattr(tree, "pericolo_rami", None))
    _set(ws, "E27", getattr(tree, "pericolo_tronco", None))
    _set(ws, "E28", getattr(tree, "pericolo_colletto", None))
    _set(ws, "E29", getattr(tree, "pericolo_zolla", None))

    # --- BERSAGLIO E IMPULSO (classi dal DB) ---
    _set(ws, "H32", getattr(tree, "moltiplicatore", None))
    _set(ws, "G32", getattr(tree, "bersaglio_chioma", None))  # classe albero/chioma
    _set(ws, "R32", getattr(tree, "bersaglio_ramo", None))    # classe rami
    _set(ws, "B33", getattr(tree, "bersaglio_chioma_value", None)
         or getattr(tree, "bersaglio_chioma_tipo", None))
    _set(ws, "M33", getattr(tree, "bersaglio_ramo_value", None)
         or getattr(tree, "bersaglio_ramo_tipo", None))

    # --- RISULTATI DEI CALCOLI - PRE INTERVENTO (dal JSON rischio in DB) ---
    # Impulso (I): momento in kgm/s e classe, per chioma/albero e per rami.
    _set(ws, "J33", _round1(attuale.get("crown_momentum_kgms")))
    _set(ws, "L33", attuale.get("crown_impulso_class"))
    _set(ws, "S33", _round1(attuale.get("branch_momentum_kgms")))
    _set(ws, "T33", attuale.get("branch_impulso_class"))
    # Pericolo (classe) per componente
    _set(ws, "E35", _comp(attuale, "rami").get("pericolo_class"))
    _set(ws, "E36", _comp(attuale, "tronco").get("pericolo_class"))
    _set(ws, "E37", _comp(attuale, "colletto").get("pericolo_class"))
    _set(ws, "E38", _comp(attuale, "zolla").get("pericolo_class"))
    # Rischio (R): rapporto + descrizione
    _set(ws, "I35", _risk_cell(attuale.get("rami")))
    _set(ws, "I36", _risk_cell(attuale.get("tronco")))
    _set(ws, "I37", _risk_cell(attuale.get("colletto")))
    _set(ws, "I38", _risk_cell(attuale.get("zolla")))

    # --- PRESCRIZIONI (righe 42-44) ---
    _write_column_block(ws, "B", 42, 44, _as_list(getattr(tree, "prescrizioni_col", None)))
    _write_column_block(ws, "L", 42, 44, _as_list(getattr(tree, "prescrizioni_mit", None)))
    _write_column_block(ws, "Q", 42, 44, _as_list(getattr(tree, "prescrizioni_val", None)))
    _set(ws, "J42", getattr(tree, "urgenza", None))
    _set(ws, "T42", getattr(tree, "monitoraggio", None))

    # --- POST-INTERVENTO: dimensioni + rischio residuo ---
    _set(ws, "C47", getattr(tree, "post_tree_height_m", None))
    _set(ws, "J47", getattr(tree, "post_branch_diam_cm", None))
    _set(ws, "M47", getattr(tree, "post_branch_length_m", None))
    _set(ws, "O47", getattr(tree, "post_branch_height_m", None))
    _set(ws, "T47", getattr(tree, "post_target_height_m", None))
    # Bersaglio (classe) post
    _set(ws, "G51", getattr(tree, "bersaglio_chioma", None))
    _set(ws, "R51", getattr(tree, "bersaglio_ramo", None))
    # Impulso (I) post
    _set(ws, "J52", _round1(residuo.get("crown_momentum_kgms")))
    _set(ws, "L52", residuo.get("crown_impulso_class"))
    _set(ws, "S52", _round1(residuo.get("branch_momentum_kgms")))
    _set(ws, "T52", residuo.get("branch_impulso_class"))
    # Pericolo (classe) post per componente
    _set(ws, "E54", _comp(residuo, "rami").get("pericolo_class"))
    _set(ws, "E55", _comp(residuo, "tronco").get("pericolo_class"))
    _set(ws, "E56", _comp(residuo, "colletto").get("pericolo_class"))
    _set(ws, "E57", _comp(residuo, "zolla").get("pericolo_class"))
    # Rischio residuo (R)
    _set(ws, "I54", _risk_cell(residuo.get("rami")))
    _set(ws, "I55", _risk_cell(residuo.get("tronco")))
    _set(ws, "I56", _risk_cell(residuo.get("colletto")))
    _set(ws, "I57", _risk_cell(residuo.get("zolla")))

    # --- NOTE ---
    _set(ws, "C59", getattr(tree, "comments", None))


# Ultima colonna da tenere: V = 22. Tutto ciò che sta dopo (liste di lookup e
# blocchi di calcolo — la "descrizione") viene eliminato.
LAST_COL = 22


def _trim_and_clean(ws) -> None:
    """Elimina tutte le colonne dopo la V e ripulisce il foglio.

    - toglie i merge nella zona a destra di V;
    - rimuove le convalide dati (i menu a tendina puntavano alle liste tagliate);
    - cancella le colonne da W in poi;
    - azzera le formule rimaste (punterebbero a celle ormai eliminate): la scheda
      resta con le sole etichette e i valori presi dal database.
    """
    # 1) unmerge di tutti i range che toccano le colonne oltre V
    for m in [r for r in ws.merged_cells.ranges if r.max_col > LAST_COL]:
        ws.unmerge_cells(str(m))

    # 2) via le convalide dati (dropdown verso liste che stiamo per cancellare)
    ws.data_validations.dataValidation = []

    # 3) elimina fisicamente le colonne da W (23) fino all'ultima
    if ws.max_column > LAST_COL:
        ws.delete_cols(LAST_COL + 1, ws.max_column - LAST_COL)

    # 4) rimuove le formule residue (niente più riferimenti alla parte tagliata)
    for row in ws.iter_rows(min_col=1, max_col=LAST_COL):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.value = None

    # 5) via le immagini del template (illustrazioni/loghi della "descrizione");
    #    il logo SilvaePro viene poi aggiunto da _add_logo.
    ws._images = []


def _set(ws, coord: str, value) -> None:
    """Scrive un valore solo se non nullo/non vuoto (evita di cancellare il
    template con dei ``None``). Le stringhe vuote vengono ignorate."""
    if value is None:
        return
    if isinstance(value, str) and value.strip() == "":
        return
    ws[coord] = value


# --- API pubblica --------------------------------------------------------------

def _safe_name(tree) -> str:
    raw = str(getattr(tree, "custom_id", None) or getattr(tree, "id", "scheda"))
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in raw)


def build_scheda_xlsx(tree) -> bytes:
    """Genera i byte di un file .xlsx (scheda ORD) per un singolo albero."""
    wb = openpyxl.load_workbook(io.BytesIO(_template_bytes()))  # keep_vba=False → .xlsx
    ws = wb[SHEET_NAME]
    _fill_ord_sheet(ws, tree)
    _trim_and_clean(ws)
    _setup_id_cell(ws, getattr(tree, "custom_id", None))
    _add_logos(ws)
    # I fogli di lookup (A, TRG-P, TRG-S) sono solo "descrizione": via anche quelli
    for name in list(wb.sheetnames):
        if name != SHEET_NAME:
            wb.remove(wb[name])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _convert_dir_to_pdf(src_dir: str, out_dir: str) -> None:
    """Converte tutti gli .xlsx di ``src_dir`` in PDF dentro ``out_dir`` con
    LibreOffice (una sola invocazione per tutti i file). Se LibreOffice non è
    disponibile o la conversione fallisce, ``out_dir`` resta senza PDF."""
    if not SOFFICE_BIN:
        return
    files = [os.path.join(src_dir, f) for f in os.listdir(src_dir) if f.endswith(".xlsx")]
    if not files:
        return
    os.makedirs(out_dir, exist_ok=True)
    profile = os.path.join(src_dir, "_lo_profile")
    try:
        subprocess.run(
            [SOFFICE_BIN, "--headless",
             f"-env:UserInstallation=file://{profile}",
             "--convert-to", "pdf", "--outdir", out_dir, *files],
            check=False, capture_output=True, timeout=300,
        )
    except (subprocess.SubprocessError, OSError):
        pass


def build_schede(trees) -> tuple[str, bytes, str]:
    """Genera le schede degli alberi dati e le impacchetta in uno ``.zip`` con
    due cartelle: ``excel/`` (gli .xlsx) e ``pdf/`` (i PDF corrispondenti).

    Ritorna ``(filename, content, mimetype)``.
    """
    trees = list(trees)
    os.makedirs(WORK_BASE, exist_ok=True)
    workdir = tempfile.mkdtemp(prefix="schede_", dir=WORK_BASE)
    try:
        excel_dir = os.path.join(workdir, "excel")
        os.makedirs(excel_dir)

        seen: dict[str, int] = {}
        for t in trees:
            name = _safe_name(t)
            seen[name] = seen.get(name, 0) + 1
            suffix = "" if seen[name] == 1 else f"_{seen[name]}"
            with open(os.path.join(excel_dir, f"scheda_{name}{suffix}.xlsx"), "wb") as fh:
                fh.write(build_scheda_xlsx(t))

        pdf_dir = os.path.join(workdir, "pdf")
        _convert_dir_to_pdf(excel_dir, pdf_dir)

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for f in sorted(os.listdir(excel_dir)):
                if f.endswith(".xlsx"):
                    zf.write(os.path.join(excel_dir, f), f"excel/{f}")
            if os.path.isdir(pdf_dir):
                for f in sorted(os.listdir(pdf_dir)):
                    if f.endswith(".pdf"):
                        zf.write(os.path.join(pdf_dir, f), f"pdf/{f}")
        return "schede_albero.zip", buf.getvalue(), ZIP_MIME
    finally:
        shutil.rmtree(workdir, ignore_errors=True)
